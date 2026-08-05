"""
BANNER CAPABILITY MAP  ->  Excel + an interactive one-page presentation

The negative space: what does your institution already OWN in Banner, and not use?

Departments keep asking to BUY
third-party products for things Banner already does. This turns that from an opinion
into evidence. For every capability it asks Banner three questions Banner itself can
answer, and none of them is an Argos report:

    1. Do the SCREENS ship and exist here?     (GUBOBJS, object type FORM)
    2. Can anyone even reach them?             (GUVUACC, the Banner security view)
    3. Is there any DATA behind them, recent?  (COUNT + MAX activity_date, plus the
                                                whole table family from statistics)

'Forms exist + people hold the keys + zero data' = owned, wired up, never used.
That is the whole argument, and it is measured, not asserted.

    python bcm_report.py --out capability_map.xlsx
    python bcm_report.py --out capability_map.xlsx --html capability_map.html

The .html is self-contained (no internet needed): click any capability on the left
and its screens, tables, record counts and user counts fill the panel on the right.

WHERE THIS CONNECTS
-------------------
a read-only Banner database. Every statement is a SELECT. Credentials come
from .env (never committed). Standard Oracle thick-mode client setup.

THE ONE CAVEAT THAT KEEPS IT HONEST
-----------------------------------
Empty is not the same as unlicensed. A capability your institution never paid for would also be
empty. This measures USE, not entitlement. It is written on the Method sheet and in
the page footer, because the credibility of the whole thing depends on not overstating.
"""
#  GENERATED FILE. Do not edit.
#  Produced by port_to_public.py from the internal engine; hand edits are
#  overwritten on the next port. Change the engine, or the public catalog in
#  _catalog_public.py, and regenerate.
import argparse
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import time
from datetime import datetime

try:
    import oracledb
    from dotenv import load_dotenv
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as e:
    sys.exit(f"Missing a package: {e.name}\n\n    pip install oracledb python-dotenv openpyxl\n")


# ---------------------------------------------------------------------------------
#  THE CATALOG.  Curated, not scanned.
#
#    forms    : LIKE pattern for the screens (proves the capability is installed)
#    forms_desc: match screens by DESCRIPTION instead, when no clean prefix exists
#    owner/tbl : the canonical data table (the judge of use; real COUNT + activity)
#    tbls     : the table FAMILY to show in the drill-down. Items ending in % are
#               resolved to every matching table from Banner statistics; plain names
#               are looked up one by one. Defaults to just the canonical table.
#    buys     : what a department buys instead, when they do
#    context  : set for capabilities whose emptiness is EXPECTED and NOT a gap
# ---------------------------------------------------------------------------------
CATALOG = [
    dict(name="Fixed Assets", buys="asset-inventory system",
         forms="FFA%", owner="FIMSMGR", tbl="FFBMAST", tbls=["FIMSMGR.FF%"],
         note="The asset master (FFBMAST) holds tracked, depreciated assets. Banner auto-flags "
              "asset candidates from purchasing into FFBOTAG; if those are never converted into "
              "FFBMAST records, asset accounting is not being done even though origination tags exist."),
    dict(name="Event Management", buys="EMS / 25Live (events, room scheduling)",
         forms="SLA%", owner="SATURN", tbl="SLBEVNT", tbls=["SATURN.SLB%"],
         note="Banner's own event and room-booking module. If SLBEVNT is empty or last touched "
              "years ago, event booking is handled elsewhere or not at all. Buildings and rooms may "
              "still be defined for class scheduling, so read the SLBEVNT row as the judge."),
    dict(name="HR Recruitment / Applicant Tracking", buys="NEOGOV / PageUp / Cornerstone",
         forms="PAAAPPL", owner="PAYROLL", tbl="PABAPPL", tbls=["PAYROLL.PAB%"],
         note="Applicant, requisition and faculty-applicant tables. If empty, recruiting is done "
              "outside Banner."),
    dict(name="Employee Skills Inventory", buys="skills / competency platform",
         forms="PPA%", owner="PAYROLL", tbl="PPRSKIL", tbls=["PAYROLL.PPRSKIL", "PAYROLL.PPRCERT"],
         note="The employee skills matrix (PPRSKIL). Related tracking (certifications, honors) may "
              "be used even when the skills matrix is empty, so read this one narrowly."),
    dict(name="Effort Certification", buys="effort-certification software",
         forms_desc="%EFFORT%", owner="PAYROLL", tbl="PTRECPD",
         note="Matched by form description, not prefix, so the count is effort-cert forms only. If "
              "your institution has grants, effort reporting may be a compliance requirement; verify "
              "the obligation before calling an empty result a gap."),
    #  Banner Workflow does not fit the form/table shape: it is a separate engine with
    #  its own schema, its own users, and its own runtime tables. Measured specially
    #  (see measure_workflow). Its verdict is computed from the real processes defined.
    dict(name="Banner Workflow", buys="Kuali / DocuSign / Formstack (approvals, e-forms)",
         special="workflow", owner="WORKFLOW", tbl="ENG_WORKFLOW",
         lf_stat="processes built", lf_sec="Business processes ever defined in Banner Workflow",
         note="Banner Workflow is a separate engine that automates approvals and business "
              "processes. Every fresh install ships with one sample process named 'system "
              "verification', the built-in installation test. If that is the only process ever "
              "defined, the engine was stood up and never used for real work. EPAF is a separate "
              "mechanism from this engine."),
    dict(name="Salary Planner (budget planning)", buys="budget-planning software (e.g. Axiom)",
         forms_desc="%SALARY PLANNER%", owner="POSNCTL", tbl="NBREHDR",
         tbls=["POSNCTL.NBREHDR", "POSNCTL.NBREPSA", "POSNCTL.NBREJLD", "POSNCTL.NTRSPEX"],
         note="Banner's native Salary Planner builds salary and position-budget scenarios. If its "
              "extract tables (NBREHDR, NBREPSA) are empty, the native worksheet flow is not used. "
              "Some institutions build a custom table for this instead; if native is empty, check "
              "whether a custom process fills the role before concluding anything."),
    dict(name="Purchasing / e-Procurement", buys="Jaggaer / Unimarket",
         forms="FPA%", owner="FIMSMGR", tbl="FPBREQH", tbls=["FIMSMGR.FP%"],
         note="Requisitions and purchase orders. High volumes here mean e-procurement runs in Banner."),
    dict(name="EPAF Personnel e-Forms", buys="DocuSign / Kuali workflow",
         forms="NOA%", owner="POSNCTL", tbl="NOBTRAN", tbls=["POSNCTL.NOB%"],
         note="Electronic Personnel Action Forms. Activity here means e-forms for personnel actions "
              "already run in Banner."),
    dict(name="Web Time Entry", buys="timekeeping product",
         forms="PHA%", owner="PAYROLL", tbl="PHRHOUR",
         tbls=["PAYROLL.PHRHOUR", "PAYROLL.PHREARN", "PAYROLL.PHRELPR"],
         note="Employee time entry. High volumes mean timekeeping runs in Banner."),
    dict(name="Benefits / Open Enrollment", buys="benefits-administration platform",
         forms="PDA%", owner="PAYROLL", tbl="PDRBCOV", tbls=["PAYROLL.PDR%"],
         note="Benefit coverage and enrollment."),
    dict(name="Grants / Research Accounting", buys="grants-management system",
         forms="FRA%", owner="FIMSMGR", tbl="FRBGRNT",
         tbls=["FIMSMGR.FRBGRNT", "FIMSMGR.FRRGRNT", "FIMSMGR.FRRBUDG"],
         note="The grant ledger for sponsored programs."),
    dict(name="Communication Management (BCM)", buys="Slate / Constant Contact / mass email",
         forms="GCA%", owner="GENERAL", tbl="GCRLENT", tbls=["GENERAL.GC%"],
         note="Banner's built-in mass-communication engine. Large volumes in the GC* tables mean a "
              "mass-comms engine already runs in Banner."),
    dict(name="Population Selection / Letter Gen", buys="mail-merge / CRM segmentation",
         forms="GLA%", owner="GENERAL", tbl="GLBEXTR",
         tbls=["GENERAL.GLBEXTR", "GENERAL.GLRSLCT"],
         note="Audience extracts and letter generation."),
    dict(name="Advancement / Alumni / Fundraising", buys="Raiser's Edge / fundraising CRM",
         forms="APA%", owner="ALUMMGR", tbl="APBCONS",
         note="Alumni and fundraising. If the ALUMMGR schema / APBCONS table is absent, the module "
              "is not installed here, which may be a licensing choice rather than shelfware."),
    dict(name="Housing / Residence Life", buys="housing-management system",
         forms="SLR%", owner="SATURN", tbl="SLRRASG", context="expected-absent",
         tbls=["SATURN.SLR%"],
         note="Room assignments. If your institution has no residence halls, an empty result is "
              "expected here, not a finding. Included so its absence is never mistaken for a gap."),
]


RECENT_MONTHS = 18      # activity within this many months = in use
STALE_MONTHS = 36       # older than this = abandoned

#  Service accounts hold access no human uses, and counting them inflates every
#  "people who can reach it" number on the map. The prefix is the one genuinely
#  institution-specific string in the SQL, so it is configuration, not a hard-coded
#  literal: that keeps the public build and this one the same code.
#  Set this, or BCM_SERVICE_ACCOUNT_PREFIX in .env, to the prefix your non-human
#  Banner accounts share. Left empty, service accounts are counted as people and
#  every "can reach it" number on the map runs high.
SVC_PREFIX = ""


def svc_filter():
    """Resolved on every call, never once at import.

    This was a bug worth keeping the comment for: read at import time, the value is
    fixed BEFORE connect() calls load_dotenv(), so setting the prefix in .env had no
    effect and every access count on the map came out inflated. Nothing failed, nothing
    warned, the report was just quietly wrong. Late binding is what makes .env work."""
    p = os.getenv("BCM_SERVICE_ACCOUNT_PREFIX", SVC_PREFIX)
    return (f"AND guvuacc_user NOT LIKE '{p}%'" if p else
            "/* no service-account prefix configured: service accounts are counted */")


def init_oracle():
    try:
        oracledb.init_oracle_client(); return
    except Exception:
        pass
    for d in [os.getenv("ORACLE_CLIENT_DIR"), r"C:\oracle\instantclient",
              r"C:\oracle\instantclient_21_0", r"C:\instantclient"]:
        if not d:
            continue
        try:
            oracledb.init_oracle_client(lib_dir=d); return
        except Exception:
            continue
    sys.exit("\n  Could not load the Oracle Instant Client. If another Oracle tool\n"
             "  (SQL Developer, etc.) works on this machine\n"
             "  the client is installed: put the folder holding oci.dll in .env as\n"
             "  ORACLE_CLIENT_DIR=<folder>.\n")


def connect():
    load_dotenv()
    miss = [k for k in ("BANNER_HOST", "BANNER_PORT", "BANNER_SERVICE",
                        "BANNER_USER", "BANNER_PASS") if not os.getenv(k)]
    if miss:
        sys.exit(f"Missing from .env: {', '.join(miss)}")
    init_oracle()
    dsn = oracledb.makedsn(os.environ["BANNER_HOST"], int(os.environ["BANNER_PORT"]),
                           service_name=os.environ["BANNER_SERVICE"])
    #  RETRY. The first connect in a fresh scheduler worker can fail on a cold
    #  connection, and here that is worse than a lost row. A failed connection makes
    #  tables unreadable, unreadable tables read as empty, and empty reads as
    #  "Owned, not used". A network hiccup could print a live capability as
    #  shelfware in front of leadership. Three tries, 2s apart.
    last = None
    for attempt in range(1, 4):
        try:
            con = oracledb.connect(user=os.environ["BANNER_USER"],
                                   password=os.environ["BANNER_PASS"], dsn=dsn)
            break
        except Exception as e:
            last = e
            if attempt < 3:
                print(f"  connection attempt {attempt} failed, retrying in 2s...")
                time.sleep(2)
    else:
        sys.exit(f"\n  Could not connect after 3 attempts: {last}\n")
    print(f"  connected to {os.environ['BANNER_SERVICE']} as {con.username}  (read-only)")
    return con


def one(cur, sql, args=None):
    cur.execute(sql, args or {})
    r = cur.fetchone()
    return r[0] if r else None


def form_names(cur, e):
    """The screens for a capability: (name, description). Prefix or description match."""
    if e.get("forms_desc"):
        cur.execute("""SELECT gubobjs_name, gubobjs_desc FROM general.gubobjs
                        WHERE gubobjs_objt_code='FORM' AND UPPER(gubobjs_desc) LIKE :d
                        ORDER BY gubobjs_name""", d=e["forms_desc"])
    else:
        cur.execute("""SELECT gubobjs_name, gubobjs_desc FROM general.gubobjs
                        WHERE gubobjs_objt_code='FORM' AND gubobjs_name LIKE :p
                        ORDER BY gubobjs_name""", p=e["forms"])
    return [(n, d) for n, d in cur.fetchall()]


def access_count(cur, e, names):
    if e.get("forms_desc"):
        if not names:
            return 0
        inlist = ",".join(f"'{n}'" for n, _ in names)   # Banner object codes
        return one(cur, f"""SELECT COUNT(DISTINCT guvuacc_user) FROM bansecr.guvuacc
                             WHERE guvuacc_object IN ({inlist})
                               {svc_filter()}
                               AND (guvuacc_role IS NULL OR guvuacc_role NOT IN
                                   ('BAN_DEFAULT_NO_ACCESS','BAN_DEFAULT_CONNECT'))""") or 0
    return one(cur, f"""SELECT COUNT(DISTINCT guvuacc_user) FROM bansecr.guvuacc
                         WHERE guvuacc_object LIKE :p {svc_filter()}
                           AND (guvuacc_role IS NULL OR guvuacc_role NOT IN
                               ('BAN_DEFAULT_NO_ACCESS','BAN_DEFAULT_CONNECT'))""",
               {"p": e["forms"]}) or 0


def _object_where(e, names):
    """The WHERE fragment that selects this capability's screens inside GUVUACC.

    Two shapes, because the catalog matches screens two ways. Returned as literal SQL
    (never user input: these are catalog constants and Banner object codes) so the same
    text can go both to the database and into the SQL drawer the audience reads.
    """
    if e.get("forms_desc"):
        if not names:
            return None
        inlist = ",".join(f"'{n}'" for n, _ in names)
        return f"guvuacc_object IN ({inlist})"
    return f"guvuacc_object LIKE '{e['forms']}'"


#  Banner hands out access four ways, and GUVUACC names the vehicle on every row:
#  guvuacc_group (a security GROUP, a bundle of classes), guvuacc_class (a security
#  CLASS, a bundle of screens), or neither, which means it was pinned straight onto
#  the person. The map used to collapse all of that into one headcount. This reads the
#  vehicle instead, so the drill-down stops at the group and never names an individual.
#  Banner names its groups after the JOB, not after the system: GTVSGRP_DESC reads
#  'Admissions Manager', 'IT Chief Information and Security Officer'. Joining it in is
#  what turns a list of codes into something a director can read without a translator.
#  Classes carry GTVCLAS_COMMENTS instead (there is no GTVCLAS_DESC column).
#
#  Oracle has no positional GROUP BY, so every expression is repeated below rather than
#  written as GROUP BY 1,2. The WITH clause keeps that from becoming unreadable, and it
#  means the query shown in the SQL drawer is the query that runs.
GROUP_PATHS = """
WITH acc AS (
    SELECT CASE WHEN guvuacc_group IS NOT NULL THEN 'Group'
                WHEN guvuacc_class IS NOT NULL THEN 'Class'
                ELSE 'Direct to person' END                             AS kind,
           COALESCE(guvuacc_group, guvuacc_class, '(no group or class)') AS path_name,
           guvuacc_object, guvuacc_role, guvuacc_user
      FROM bansecr.guvuacc
     WHERE {obj}
       {svc}
       AND (guvuacc_role IS NULL OR guvuacc_role NOT IN
            ('BAN_DEFAULT_NO_ACCESS','BAN_DEFAULT_CONNECT'))
)
SELECT a.kind,
       a.path_name,
       COALESCE(g.gtvsgrp_desc, c.gtvclas_comments)               AS descr,
       COUNT(DISTINCT a.guvuacc_object)                           AS screens,
       COUNT(DISTINCT CASE WHEN a.guvuacc_role LIKE '%\\_M' ESCAPE '\\'
                           THEN a.guvuacc_object END)             AS can_change,
       COUNT(DISTINCT a.guvuacc_user)                             AS people
  FROM acc a
  LEFT JOIN bansecr.gtvsgrp g ON g.gtvsgrp_code       = a.path_name AND a.kind = 'Group'
  LEFT JOIN bansecr.gtvclas c ON c.gtvclas_class_code = a.path_name AND a.kind = 'Class'
 GROUP BY a.kind, a.path_name, COALESCE(g.gtvsgrp_desc, c.gtvclas_comments)
 ORDER BY a.kind, people DESC, a.path_name
"""

#  Groups first, then classes, then whatever was pinned directly onto a person. That
#  last bucket is the one worth staring at: it is access that no group governs.
KIND_ORDER = {"Group": 0, "Class": 1, "Direct to person": 2}


def group_paths(cur, e, names):
    """Which security groups and classes carry this capability, and how wide each one is.

    One row per grant vehicle. `people` is a headcount, never a list of names: the
    boss asked for the group level and stopping there is the whole point, not a
    limitation we ran into.
    """
    obj = _object_where(e, names)
    if obj is None:
        return []
    try:
        cur.execute(GROUP_PATHS.format(obj=obj, svc=svc_filter()))
    except Exception:
        return []
    out = [dict(kind=k, name=n, desc=(d or "").strip(), screens=s, change=c or 0, people=p)
           for k, n, d, s, c, p in cur.fetchall()]
    out.sort(key=lambda d: (KIND_ORDER.get(str(d["kind"]), 9), -int(d["people"]), str(d["name"])))
    return out


def table_family(cur, e):
    """The associated tables, with row estimates from Banner statistics (instant, and
    honest about being estimates). Prefix specs ('FIMSMGR.FF%') expand to the family.

    Two rules keep the panel honest:
      - Drop validation tables (Banner names them with 'V' in the third position, e.g.
        FFVACON). They carry a handful of seed codes and only add noise.
      - The canonical table (the one the verdict is judged on) is ALWAYS shown and
        pinned first, even when it is empty and would otherwise sort to the bottom or
        fall off the list. The whole argument rests on that row; it cannot go missing.
    """
    specs = e.get("tbls") or [f"{e['owner']}.{e['tbl']}"]
    canon = f"{e['owner']}.{e['tbl']}"
    seen, out = set(), []
    #  Banner writes a plain-English description of every table into DBA_TAB_COMMENTS
    #  (FFBOTAG -> 'Fixed Asset Origination Tag Table'). Join it in so the panel can say
    #  what each table IS, not just its code.
    base = """SELECT t.owner, t.table_name, t.num_rows,
                     TO_CHAR(t.last_analyzed,'YYYY-MM-DD'), c.comments
                FROM dba_tables t
                LEFT JOIN dba_tab_comments c
                       ON c.owner=t.owner AND c.table_name=t.table_name
               WHERE t.owner=:o AND {cond}"""
    for spec in specs:
        owner, frag = spec.split(".", 1)
        if frag.endswith("%"):
            cur.execute(base.format(cond="t.table_name LIKE :f") +
                        " ORDER BY t.num_rows DESC NULLS LAST FETCH FIRST 25 ROWS ONLY",
                        o=owner, f=frag)
        else:
            cur.execute(base.format(cond="t.table_name=:f"), o=owner, f=frag)
        for o, t, n, la, cm in cur.fetchall():
            key = f"{o}.{t}"
            if key in seen:
                continue
            if len(t) >= 3 and t[2] == "V" and key != canon:   # skip validation tables
                continue
            seen.add(key)
            out.append({"t": key, "rows": None if n is None else int(n),
                        "analyzed": la, "primary": key == canon,
                        "desc": (cm or "").strip()})
    out.sort(key=lambda r: (0 if r["primary"] else 1,
                            -(r["rows"] if r["rows"] is not None else -1)))

    #  Census across the WHOLE module, not just what fits the panel. For a prefix spec
    #  this is the real size of the module and how little of it holds data, which is the
    #  point: '41 tables in this module, 2 hold any data' lands harder than a short list.
    total = with_data = 0
    for spec in specs:
        owner, frag = spec.split(".", 1)
        if not frag.endswith("%"):
            continue                     # census is a MODULE scan; skip hand-picked tables
        cur.execute(f"""SELECT COUNT(*), COUNT(CASE WHEN num_rows>0 THEN 1 END)
                          FROM dba_tables
                         WHERE owner=:o AND table_name LIKE :f
                           AND NOT (LENGTH(table_name)>=3 AND SUBSTR(table_name,3,1)='V')""",
                    o=owner, f=frag)
        tt, dd = cur.fetchone()
        total += tt or 0
        with_data += dd or 0
    return out[:12], {"total": total, "with_data": with_data}


#  Friendly, factual labels for the Workflow runtime tables (their DBA_TAB_COMMENTS are
#  blank). These describe what each table HOLDS; they make no claim.
WF_DESC = {
    "WORKFLOW.ENG_WORKFLOW": "Workflow instances that have ever run",
    "WORKFLOW.ENG_WORKITEM": "Individual work items (tasks) across all workflows",
    "WORKFLOW.ENG_COMPLETED_EVENTS": "Completed workflow events",
    "WORKFLOW.PROCESS_DEFINITION": "Business processes defined in the Workflow designer",
    "WORKFLOW.WFUSER": "Workflow user accounts",
}


def measure_workflow(cur):
    """Banner Workflow, measured on its own terms: processes built, users, instances run."""
    def c(sql):
        try:
            return one(cur, sql)
        except Exception:
            return None
    #  The processes ever DEFINED become the 'forms' list: name + description. Anyone can
    #  read whether these are real business processes or the shipped install test.
    forms = []
    try:
        cur.execute("SELECT name, description FROM workflow.process_definition ORDER BY name")
        forms = [(n, d) for n, d in cur.fetchall()]
    except Exception:
        pass
    access = c("SELECT COUNT(*) FROM workflow.wfuser") or 0
    rows = c("SELECT COUNT(*) FROM workflow.eng_workflow")           # instances ever run
    tables = []
    for t in ["ENG_WORKFLOW", "PROCESS_DEFINITION", "ENG_WORKITEM", "ENG_COMPLETED_EVENTS",
              "WFUSER"]:
        n = c(f"SELECT COUNT(*) FROM workflow.{t}")
        key = f"WORKFLOW.{t}"
        tables.append({"t": key, "rows": n, "analyzed": None,
                       "primary": t == "ENG_WORKFLOW", "desc": WF_DESC.get(key, "")})
    census = {"total": c("SELECT COUNT(*) FROM dba_tables WHERE owner='WORKFLOW'") or 0,
              "with_data": c("SELECT COUNT(*) FROM dba_tables WHERE owner='WORKFLOW' "
                             "AND num_rows>0") or 0}
    return forms, access, rows, None, tables, census, False


def max_activity(cur, owner, tbl):
    """Newest activity date, WITHOUT hard-coding <tbl>_ACTIVITY_DATE. Banner's naming
    convention is common but not universal (local, work and interface tables break it),
    so discover the real column from the catalog first. Returns None if the table has
    no *_ACTIVITY_DATE column at all, rather than a misleading query error."""
    try:
        col = one(cur, r"""SELECT column_name FROM all_tab_columns
                            WHERE owner=:o AND table_name=:t
                              AND column_name LIKE '%\_ACTIVITY\_DATE' ESCAPE '\'
                            ORDER BY column_name FETCH FIRST 1 ROWS ONLY""",
                  {"o": owner, "t": tbl})
        if not col:
            return None
        return one(cur, f"SELECT TO_CHAR(MAX({col}),'YYYY-MM-DD') FROM {owner}.{tbl}")
    except Exception:
        return None


def measure(cur, e):
    forms = form_names(cur, e)
    access = access_count(cur, e, forms)
    rows, err = None, False
    try:
        rows = one(cur, f"SELECT COUNT(*) FROM {e['owner']}.{e['tbl']}")
    except Exception as ex:
        #  ORA-00942 (table or view does not exist) is the only error that means
        #  'Not installed'. A timeout, a privilege problem, or VPD hiding rows is a
        #  READ FAILURE and must never be reported as a capability nobody uses.
        err = "ORA-00942" not in str(ex).upper()
    last = None
    if rows:
        last = max_activity(cur, e["owner"], e["tbl"])
    tables, census = table_family(cur, e)
    #  If stats never listed the canonical table but it really exists, put it in front
    #  with the EXACT count we just measured. The judge row is never allowed to vanish.
    canon = f"{e['owner']}.{e['tbl']}"
    if rows is not None and not any(t["t"] == canon for t in tables):
        cm = ""
        try:
            cm = (one(cur, "SELECT comments FROM dba_tab_comments WHERE owner=:o AND table_name=:t",
                      {"o": e["owner"], "t": e["tbl"]}) or "").strip()
        except Exception:
            pass
        tables.insert(0, {"t": canon, "rows": rows, "analyzed": last,
                          "primary": True, "desc": cm})
    return forms, access, rows, last, tables, census, err


def months_since(datestr):
    if not datestr:
        return None
    d = datetime.strptime(datestr, "%Y-%m-%d")
    return (datetime.now() - d).days / 30.0


def verdict(e, rows, last, error=False):
    if e.get("context") == "expected-absent":
        return "Expected absent", "not a gap"
    if error:
        #  The table is there but we could not read it (timeout, privilege, row-level
        #  security). That is a read failure, NEVER a use signal in either direction.
        return "Unable to verify", "the table could not be read (timeout or privilege), not a use signal"
    if rows is None:
        return "Not installed", "no Banner table for this capability is present here"
    if rows == 0:
        return "Owned, not used", ("the table Banner ships for this holds no rows; "
                                   "confirm no self-service path writes elsewhere")
    m = months_since(last)
    if m is None:
        return "Owned, data present", "data exists but the table carries no activity date to age it"
    if m <= RECENT_MONTHS:
        return "In use", "recent activity on the canonical table"
    if m >= STALE_MONTHS:
        return "Abandoned", f"newest activity dates to {last[:4] if last else '?'}"
    return "Fading", "activity present but slowing"


ORDER = {"Owned, not used": 0, "Abandoned": 1, "Custom-built": 2, "Not installed": 3,
         "Fading": 4, "Owned, data present": 5, "In use": 6, "Expected absent": 7,
         "Unable to verify": 8}
FILL = {"Owned, not used": "F8CBAD", "Abandoned": "F8CBAD", "Custom-built": "FFF2CC",
        "Fading": "FFF2CC", "Not installed": "E7E6E6", "In use": "C6E0B4",
        "Owned, data present": "C6E0B4", "Expected absent": "D9D9D9",
        "Unable to verify": "FFE699"}


def build_rows(cur):
    out = []
    for e in CATALOG:
        if e.get("special") == "workflow":
            forms, access, rows, last, tables, census, err = measure_workflow(cur)
            #  Workflow is not governed through Banner security groups; it carries its
            #  own user list. Showing an empty group table is honest, inventing one is not.
            paths = []
        else:
            forms, access, rows, last, tables, census, err = measure(cur, e)
            paths = group_paths(cur, e, forms)
        if e.get("force_verdict"):
            v, why = e["force_verdict"], e.get("force_why", "")
        elif e.get("special") == "workflow":
            #  Workflow cannot be judged by a row count: every fresh install ships with
            #  one sample process named 'system verification', the vendor's own install
            #  test, and running it leaves rows behind. So the judge is the PROCESSES
            #  DEFINED, minus that sample. Computed per institution, never assumed.
            real = [f for f in forms if "system verification" not in (f[0] or "").lower()]
            v, why = (("In use", f"{len(real)} business process(es) defined") if real else
                      ("Owned, not used",
                       "only the shipped 'system verification' install test"))
        else:
            v, why = verdict(e, rows, last, err)
        #  ACCESS DEBT: the capability is dark, and a group still hands out the right to
        #  change it. Nobody is doing anything wrong; it is simply standing permission
        #  on a room no one enters, and it is the cheapest thing on this page to fix.
        debt = (v in ("Owned, not used", "Abandoned")
                and any(p["change"] for p in paths))
        out.append(dict(e=e, forms=forms, access=access, rows=rows, last=last,
                        tables=tables, census=census, verdict=v, why=why,
                        paths=paths, debt=debt))
        rs = "n/a" if rows is None else f"{rows:,}"
        ng = sum(1 for p in paths if p["kind"] == "Group")
        print(f"  {e['name']:<38} {v:<18} forms={len(forms):<3} access={access:<4} "
              f"rows={rs:<12} groups={ng}{'  <-- access debt' if debt else ''}")
    out.sort(key=lambda r: (ORDER.get(r["verdict"], 9), -(r["access"] or 0)))
    return out


# ---------------------------------------------------------------------------------
#  EXCEL
# ---------------------------------------------------------------------------------
HDR = PatternFill("solid", fgColor="1F3864")
HDRF = Font(color="FFFFFF", bold=True)


def _since_sheet(wb, changes):
    """Only appears once there is a previous run to compare against. An empty sheet on
    the first run would look like a bug; its absence is the honest signal."""
    ws = wb.create_sheet("Since last run")
    ws.append([f"What changed since {changes['prev_date']}"])
    ws["A1"].font = Font(bold=True, size=13, color="1F3864")
    ws.append([])
    ws.append(["Capability", "What changed", "Level"])
    for c in ws[3]:
        c.fill, c.font = HDR, HDRF
    for name, d in changes["by_name"].items():
        for t in d["head"]:
            ws.append([name, t, "capability"])
        for t in d["detail"]:
            ws.append([name, t, "group or class"])
    for i, w in enumerate([34, 78, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"


def build_excel(path, data, changes=None, demo=False):
    wb = Workbook(); ws = wb.active; ws.title = "Capability Map"
    if demo:
        #  A workbook travels further than a web page and loses its context on the way.
        #  The warning has to be inside the file, on the first sheet, in the first cell.
        ws.append(["DEMONSTRATION COPY, DO NOT PRESENT. The capability measurements are "
                   "real and current; the earlier run they are compared against is "
                   "INVENTED, so every 'since' number is fictional."])
        ws["A1"].font = Font(bold=True, color="FFFFFF", size=12)
        ws["A1"].fill = PatternFill("solid", fgColor="B03A2E")
        ws.append([])
    cols = ["Capability", "What departments buy for this", "Banner forms",
            "People who can reach them", "Security groups", "Direct grants",
            "Records", "Last activity", "Verdict", "Note"]
    ws.append(cols)
    hdr = ws.max_row                       # 1 normally, 3 on a demo copy
    for c in ws[hdr]:
        c.fill, c.font, c.alignment = HDR, HDRF, Alignment(horizontal="center", wrap_text=True)
    for r in data:
        e = r["e"]
        paths = r.get("paths") or []
        ngrp = sum(1 for p in paths if p["kind"] == "Group")
        ndir = sum(1 for p in paths if p["kind"] == "Direct to person")
        ws.append([e["name"], e["buys"], len(r["forms"]), r["access"], ngrp, ndir,
                   "not installed" if r["rows"] is None else r["rows"],
                   r["last"] or "-", r["verdict"], e["note"]])
        fill = PatternFill("solid", fgColor=FILL.get(r["verdict"], "FFFFFF"))
        for c in ws[ws.max_row]:
            c.fill = fill
            c.alignment = Alignment(vertical="top", wrap_text=True)
    for i, w in enumerate([34, 30, 8, 12, 10, 10, 12, 12, 16, 52], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{hdr + 1}"
    ws.auto_filter.ref = f"A{hdr}:{get_column_letter(len(cols))}{ws.max_row}"

    if changes and changes.get("by_name"):
        _since_sheet(wb, changes)
    _groups_sheet(wb, data)
    _evidence_sheet(wb, data)
    _method_sheet(wb, data)
    wb.save(path)
    print(f"\n  wrote {path}")


def _groups_sheet(wb, data):
    """The drill-down the boss asked for: capability -> the groups and classes that
    open it. One row per grant vehicle, never per person. 'People' is a headcount.

    Rows are shaded when the capability is dark and the vehicle still grants change
    rights, because that pair is the finding: standing permission on an empty room.
    """
    ws = wb.create_sheet("Access by Group")
    ws.append(["Capability", "Verdict", "Kind", "Group or class", "What that group is",
               "Screens it opens", "Of those, can change", "People holding it",
               "Standing access on a dark capability"])
    for c in ws[1]:
        c.fill, c.font, c.alignment = HDR, HDRF, Alignment(horizontal="center", wrap_text=True)
    warn = PatternFill("solid", fgColor="F8CBAD")
    for r in data:
        for p in r.get("paths") or []:
            flag = "YES" if (r.get("debt") and p["change"]) else ""
            ws.append([r["e"]["name"], r["verdict"], p["kind"], p["name"],
                       p.get("desc") or "", p["screens"], p["change"], p["people"], flag])
            if flag:
                for c in ws[ws.max_row]:
                    c.fill = warn
    for i, w in enumerate([34, 18, 16, 40, 40, 14, 16, 14, 22], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _evidence_sheet(wb, data):
    """One place that lists, per capability, the actual screens and tables measured.
    So the Excel carries the same drill-down the presentation shows."""
    ws = wb.create_sheet("Evidence")
    ws.append(["Capability", "Kind", "Object", "Detail"])
    for c in ws[1]:
        c.fill, c.font = HDR, HDRF
    for r in data:
        name = r["e"]["name"]
        for fn, fd in r["forms"]:
            ws.append([name, "screen", fn, fd])
        for t in r["tables"]:
            rc = "?" if t["rows"] is None else f"{t['rows']:,} rows (stats)"
            desc = t.get("desc") or ""
            ws.append([name, "table", t["t"],
                       f"{desc}  ({rc}, analyzed {t['analyzed'] or '-'})"])
    for i, w in enumerate([34, 8, 22, 72], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def _method_sheet(wb, data):
    ws = wb.create_sheet("Method")
    owned = [r for r in data if r["verdict"] in ("Owned, not used", "Abandoned")]
    rows = [
        ("What this is", "A map of Banner capabilities your institution owns, and whether each is used. "
         "Built to answer purchase requests for things Banner already does."),
        ("Source", "a read-only Banner database. Every statement is a SELECT."),
        ("Run by", os.getenv("USERNAME") or os.getenv("USER") or "(unknown)"),
        ("Run at", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("", ""),
        ("HOW EACH ROW IS MEASURED", ""),
        ("Banner forms", "How many screens for this capability exist here (GUBOBJS, type FORM). "
         "Proves the capability is installed, not an orphan table."),
        ("People who can reach them", "Distinct real users with access to those forms (GUVUACC). "
         "This is 'can reach', NOT 'uses'. The size of the open door, not the traffic."),
        ("Security groups / Direct grants", "How many security GROUPS open this capability, and "
         "how many grants were pinned straight onto a person with no group governing them. The "
         "full listing is on the 'Access by Group' sheet."),
        ("Records / Last activity", "Real row count and newest activity date in the canonical "
         "table. The table family on the Evidence sheet uses Banner's own statistics (estimates)."),
        ("", ""),
        ("THE GROUP LAYER  ('Access by Group' sheet)", ""),
        ("Why it stops at the group", "Banner hands out access four ways, and GUVUACC names the "
         "vehicle on every row: a security GROUP (a bundle of classes), a security CLASS (a bundle "
         "of screens), or neither, meaning it was pinned onto one person. This sheet reports the "
         "vehicle, never the individual. That is a deliberate choice, not a limit we hit: the group "
         "is the thing you can actually govern, and a per-person listing is a different review with "
         "different rules."),
        ("Screens it opens", "How many of THIS capability's screens that group or class reaches. "
         "A group usually spans several capabilities; this column counts only the ones here."),
        ("Of those, can change", "How many of them it can change rather than only view (Banner "
         "marks the maintain role with a _M suffix). This is the column that carries risk."),
        ("People holding it", "A headcount of who carries that group or class. A count, not a list."),
        ("Standing access on a dark capability", "Flagged when the capability holds no live data "
         "AND a group still grants the right to change it. Nothing is misconfigured; it is standing "
         "permission on a room nobody enters. It is the cheapest cleanup on this report, because "
         "removing it costs nothing and takes nothing away from anyone who is working."),
        ("Banner Workflow", "Shows no groups on purpose. The Workflow engine carries its own user "
         "list and is not governed through Banner security classes."),
        ("", ""),
        ("HISTORY  ('Since last run' sheet)", ""),
        ("What is kept", "Each run writes a small JSON snapshot into a history folder next to "
         "this workbook: verdict, people who can reach it, records, and every group or class "
         "by name. No individual is stored, here or anywhere else in this report."),
        ("What it is for", "One run is a photograph. Kept side by side, the runs answer a "
         "question a single run cannot: what MOVED. It is also the only way to show that a "
         "cleanup decided in a meeting actually happened."),
        ("Order of operations", "This run is compared against the previous one BEFORE it is "
         "saved. Otherwise a run becomes its own baseline and the report can never show change."),
        ("On the first run", "The 'Since last run' sheet is absent, because there is nothing "
         "to compare against yet. That absence is the honest signal, not a bug."),
        ("", ""),
        ("VERDICTS", ""),
        ("Owned, not used", "Forms ship and people hold access, but the table Banner writes to "
         "for this holds no rows. A signal worth confirming with the product owner, not a "
         "conclusion: check that no self-service path feeds the data somewhere else first."),
        ("Abandoned", "Data exists but the newest activity date is years old."),
        ("In use", "Recent activity. A department wanting a tool for this already has a working one."),
        ("Not installed", "No Banner table for this capability is present here."),
        ("Unable to verify", "The table exists but could not be read (timeout, privilege, or "
         "row-level security). Not a use signal in either direction; re-run or check the grant."),
        ("Expected absent", "Empty on purpose. Housing at a college with no residence halls. Included so its "
         "absence is never mistaken for a gap."),
        ("", ""),
        ("THE ONE CAVEAT", "Database evidence shows technical artifacts and rows visible to this "
         "read-only account. It does NOT prove license ownership, deployment, configuration, "
         "adoption, usability, or functional fit. Empty is not the same as unlicensed; a busy "
         "activity date is not the same as human use. Before telling a department 'you already "
         "own this', validate entitlement and workflow with the Banner product owner. The tool "
         "shows the door and the traffic; the contract and the fit are separate questions."),
        ("", ""),
        ("HEADLINE", f"{len(owned)} Banner capabilities are owned and idle. Each is a purchase "
         "to question before it is made."),
    ]
    for a, b in rows:
        ws.append([a, b])
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 104
    for row in ws.iter_rows():
        row[0].font = Font(bold=True)
        row[1].alignment = Alignment(wrap_text=True, vertical="top")


# ---------------------------------------------------------------------------------
#  METHOD SQL  (the exact queries behind each number, shown in the panel's SQL drawer)
#  Method only: standard Banner data-dictionary views. Runs on any institution's Banner.
# ---------------------------------------------------------------------------------
def _q_screens(e):
    if e.get("forms_desc"):
        cond = f"   AND UPPER(gubobjs_desc) LIKE '{e['forms_desc']}'"
    else:
        cond = f"   AND gubobjs_name LIKE '{e['forms']}'"
    return ("SELECT gubobjs_name, gubobjs_desc\n"
            "  FROM general.gubobjs\n"
            " WHERE gubobjs_objt_code = 'FORM'\n"
            f"{cond}\n"
            " ORDER BY gubobjs_name;")


def _q_access(e):
    obj = (" WHERE guvuacc_object IN ( :form_codes )   -- the codes from the query above\n"
           if e.get("forms_desc")
           else f" WHERE guvuacc_object LIKE '{e['forms']}'\n")
    return ("SELECT COUNT(DISTINCT guvuacc_user)\n"
            "  FROM bansecr.guvuacc\n"
            f"{obj}"
            f"   {svc_filter()}   -- service accounts are not people\n"
            "   AND (guvuacc_role IS NULL OR guvuacc_role NOT IN\n"
            "        ('BAN_DEFAULT_NO_ACCESS','BAN_DEFAULT_CONNECT'));")


def _q_groups(e):
    obj = ("guvuacc_object IN ( :form_codes )   -- the codes from the screens query"
           if e.get("forms_desc") else f"guvuacc_object LIKE '{e['forms']}'")
    return ("-- Which security groups and classes carry this capability, and how wide each\n"
            "-- one is. GUVUACC names the vehicle on every grant, so nothing has to be\n"
            "-- reconstructed: group, class, or neither (pinned straight onto a person).\n"
            "-- GTVSGRP_DESC is joined in because Banner names groups after the JOB\n"
            "-- (Admissions Manager, Payroll Supervisor), which is the readable layer.\n"
            "WITH acc AS (\n"
            "    SELECT CASE WHEN guvuacc_group IS NOT NULL THEN 'Group'\n"
            "                WHEN guvuacc_class IS NOT NULL THEN 'Class'\n"
            "                ELSE 'Direct to person' END                       AS kind,\n"
            "           COALESCE(guvuacc_group, guvuacc_class, '(none)')       AS path_name,\n"
            "           guvuacc_object, guvuacc_role, guvuacc_user\n"
            "      FROM bansecr.guvuacc\n"
            f"     WHERE {obj}\n"
            f"       {svc_filter()}\n"
            "       AND (guvuacc_role IS NULL OR guvuacc_role NOT IN\n"
            "            ('BAN_DEFAULT_NO_ACCESS','BAN_DEFAULT_CONNECT'))\n"
            ")\n"
            "SELECT a.kind,\n"
            "       a.path_name,\n"
            "       COALESCE(g.gtvsgrp_desc, c.gtvclas_comments)          AS descr,\n"
            "       COUNT(DISTINCT a.guvuacc_object)                      AS screens,\n"
            "       COUNT(DISTINCT CASE WHEN a.guvuacc_role LIKE '%\\_M' ESCAPE '\\'\n"
            "                           THEN a.guvuacc_object END)        AS can_change,\n"
            "       COUNT(DISTINCT a.guvuacc_user)                        AS people\n"
            "  FROM acc a\n"
            "  LEFT JOIN bansecr.gtvsgrp g\n"
            "         ON g.gtvsgrp_code       = a.path_name AND a.kind = 'Group'\n"
            "  LEFT JOIN bansecr.gtvclas c\n"
            "         ON c.gtvclas_class_code = a.path_name AND a.kind = 'Class'\n"
            " -- Oracle has no positional GROUP BY, so the expressions repeat here.\n"
            " GROUP BY a.kind, a.path_name, COALESCE(g.gtvsgrp_desc, c.gtvclas_comments)\n"
            " ORDER BY a.kind, people DESC, a.path_name;")


def _q_records(e):
    o, t = e["owner"], e["tbl"]
    return (f"-- how many records live in the canonical (judge) table\n"
            f"SELECT COUNT(*) FROM {o}.{t};\n\n"
            f"-- newest activity: the date column is DISCOVERED, not hard-coded\n"
            f"SELECT column_name FROM all_tab_columns\n"
            f" WHERE owner = '{o}' AND table_name = '{t}'\n"
            f"   AND column_name LIKE '%\\_ACTIVITY\\_DATE' ESCAPE '\\'\n"
            f" FETCH FIRST 1 ROWS ONLY;\n"
            f"-- then:  SELECT MAX(<that column>) FROM {o}.{t};")


def _q_tables(e):
    specs = e.get("tbls") or [f"{e['owner']}.{e['tbl']}"]
    owner, frag = specs[0].split(".", 1)
    return (f"-- the table family, row estimates from Banner's own statistics\n"
            f"SELECT table_name, num_rows, last_analyzed\n"
            f"  FROM dba_tables\n"
            f" WHERE owner = '{owner}' AND table_name LIKE '{frag}'\n"
            f" ORDER BY num_rows DESC NULLS LAST;")


def method_sql(e):
    """The four method queries for a capability, each with a title and a one-line why."""
    if e.get("special") == "workflow":
        return {
            "screens": {"t": "Which processes are built?",
                        "w": "Every business process ever defined in the Workflow designer. If the "
                             "only one is the shipped 'system verification' test, the engine was never used.",
                        "q": "SELECT name, description\n  FROM workflow.process_definition\n ORDER BY name;"},
            "access":  {"t": "Who are the workflow users?",
                        "w": "The Workflow user accounts provisioned on the engine.",
                        "q": "SELECT COUNT(*) FROM workflow.wfuser;"},
            "records": {"t": "How many instances ever ran?",
                        "w": "Workflow instances that have ever executed. This is the judge for Workflow.",
                        "q": "SELECT COUNT(*) FROM workflow.eng_workflow;"},
            "tables":  {"t": "The Workflow tables",
                        "w": "The engine's own tables and how many hold any rows.",
                        "q": "SELECT table_name, num_rows\n  FROM dba_tables\n WHERE owner = 'WORKFLOW'\n"
                             " ORDER BY num_rows DESC NULLS LAST;"},
        }
    return {
        "screens": {"t": "Do the screens exist?",
                    "w": "Banner forms that ship for this, from the object registry GUBOBJS. If they "
                         "exist, the capability is installed here, not an orphan table.",
                    "q": _q_screens(e)},
        "access":  {"t": "Can anyone reach them?",
                    "w": "Distinct users who hold access to those forms, from the security view GUVUACC. "
                         "No-access roles and service accounts are excluded. The open door, not the traffic.",
                    "q": _q_access(e)},
        "groups":  {"t": "Through which groups?",
                    "w": "The same view, read by the vehicle instead of the headcount: which security "
                         "group or class carries this capability, how many of its screens each one "
                         "opens, and how many of those it can change. Stops at the group by design.",
                    "q": _q_groups(e)},
        "records": {"t": "Is there data, and recent?",
                    "w": "A real COUNT on the canonical table, plus the newest activity date. The activity "
                         "column is discovered from the catalog, never assumed. This is the judge.",
                    "q": _q_records(e)},
        "tables":  {"t": "The table family",
                    "w": "The associated tables and their row estimates, from Banner's own optimizer "
                         "statistics (instant, and honest about being estimates).",
                    "q": _q_tables(e)},
    }


# ---------------------------------------------------------------------------------
#  HISTORY.  A single run is a photograph. Kept side by side, the runs become an
#  instrument: the map can then open with what MOVED, which is the only way to show
#  that a cleanup actually happened. The first run has nothing to compare against,
#  which is exactly why the saving has to start now and not when we want the answer.
# ---------------------------------------------------------------------------------
def verdict_cls(v):
    return v.split(",")[0].lower().replace(" ", "-")


def slug(s):
    return re.sub(r"[^a-z0-9]+", "-", str(s).lower()).strip("-")


def snapshot(data, when):
    """The small, stable shape worth keeping. Deliberately NOT the whole report: only
    the numbers a later run can meaningfully compare. Groups are stored by name so a
    grant path that appears or disappears can be named, not just counted."""
    caps = []
    for r in data:
        caps.append({
            "name": r["e"]["name"], "verdict": r["verdict"], "access": r["access"],
            "rows": r["rows"], "forms": len(r["forms"]), "debt": bool(r.get("debt")),
            "paths": {p["name"]: {"kind": p["kind"], "change": p["change"],
                                  "people": p["people"]}
                      for p in (r.get("paths") or [])},
        })
    return {"date": when, "caps": caps}


class HistoryStore:
    """Where the runs are kept. A folder on a laptop today; on a scheduler like Windmill
    the worker's filesystem is thrown away after every job, so the store has to move to
    object storage or a table.

    This is the ONLY part of the tool that touches stored history. Anything that keeps
    a dated JSON blob and can list what it holds is a valid backend: subclass, implement
    the three methods, and assign it to bcm_report.STORE. Nothing else changes, and the
    hosted version stays the same code as the laptop version rather than a fork that
    slowly drifts."""

    def __init__(self, dirpath):
        self.dirpath = dirpath

    def list(self):
        """Every stored run date, ascending, as YYYY-MM-DD strings."""
        if not os.path.isdir(self.dirpath):
            return []
        return sorted(n[:-5] for n in os.listdir(self.dirpath) if n.endswith(".json"))

    def read(self, date):
        with open(os.path.join(self.dirpath, f"{date}.json"), encoding="utf-8") as f:
            return json.load(f)

    def write(self, snap):
        os.makedirs(self.dirpath, exist_ok=True)
        p = os.path.join(self.dirpath, f"{snap['date']}.json")
        with open(p, "w", encoding="utf-8") as f:
            json.dump(snap, f, indent=1)
        return p


STORE = None                     # set by main(), or by the host before calling in


def save_snapshot(snap, dirpath=None):
    st = STORE or HistoryStore(dirpath)
    p = st.write(snap)
    print(f"  wrote {p}   (history: {len(st.list())} run(s) kept)")


def load_prev(dirpath, today):
    """The most recent run that is not today's. Re-running on the same day should not
    silently become its own baseline and report that nothing ever changes."""
    st = STORE or HistoryStore(dirpath)
    older = [d for d in st.list() if d < today]
    if not older:
        return None
    try:
        return st.read(older[-1])
    except Exception:
        return None


def list_runs(dirpath=None):
    return (STORE or HistoryStore(dirpath)).list()


def load_run(dirpath, date):
    """The stored run ON that date, or the nearest one BEFORE it. Asking for a date the
    tool did not run is normal (a meeting was in March, the run was in April), so resolve
    to the closest earlier run and say which one was used rather than refusing."""
    st = STORE or HistoryStore(dirpath)
    runs = st.list()
    older = [d for d in runs if d <= date]
    if not older:
        sys.exit(f"  No run stored on or before {date}. Runs available: "
                 + (", ".join(runs) if runs else "none yet"))
    return st.read(older[-1])


def diff_snapshots(prev, cur):
    """What moved. Split in two on purpose: `head` is what a director needs at the top
    of the page, `detail` is the group-level movement that belongs inside the card."""
    pm = {c["name"]: c for c in prev.get("caps", [])}
    by, head = {}, []
    for c in cur["caps"]:
        p = pm.get(c["name"])
        if p is None:
            by[c["name"]] = {"head": ["New on the map: not measured in the earlier run."],
                             "detail": []}
            continue
        H, D = [], []
        if p["verdict"] != c["verdict"]:
            H.append(f'Verdict moved from "{p["verdict"]}" to "{c["verdict"]}".')
        d = (c["access"] or 0) - (p["access"] or 0)
        if d:
            H.append(f'{abs(d)} {"more" if d > 0 else "fewer"} people can reach it '
                     f'({p["access"]} to {c["access"]}).')
        if p["rows"] is not None and c["rows"] is not None:
            if not p["rows"] and c["rows"]:
                H.append(f'It woke up: {c["rows"]:,} records now, none in the earlier run.')
            elif p["rows"] and not c["rows"]:
                H.append("Its records are gone: it held data before and holds none now.")
        if p.get("debt") and not c.get("debt"):
            H.append("Standing access on a dark capability is cleared.")
        elif c.get("debt") and not p.get("debt"):
            H.append("It now carries standing change access on a capability with no data.")
        pp, cp = p.get("paths") or {}, c.get("paths") or {}
        for k in sorted(set(cp) - set(pp)):
            D.append(f"New grant path: {k} ({cp[k]['people']} people).")
        for k in sorted(set(pp) - set(cp)):
            D.append(f"Grant path removed: {k} (held {pp[k]['people']} people).")
        for k in sorted(set(pp) & set(cp)):
            a, b = pp[k].get("people", 0), cp[k].get("people", 0)
            if a != b:
                D.append(f"{k}: {a} to {b} people.")
        if H or D:
            by[c["name"]] = {"head": H, "detail": D}
        head.extend((c["name"], h) for h in H)
    return {"prev_date": prev.get("date", "?"), "by_name": by, "head": head,
            "n": len(by)}


# ---------------------------------------------------------------------------------
#  TWO DATES.  Once the runs are on disk the comparison no longer needs Banner at all:
#  it is a question about two files. That matters more than it sounds. It runs off VPN,
#  it runs in a meeting, it runs in a second, and it can never disturb the database.
# ---------------------------------------------------------------------------------
def _cmp_rows(prev, cur):
    """Every capability, side by side, whether or not it moved. The ones that did not
    move are the control group: without them a reader cannot tell a stable map from a
    tool that only looked at four things."""
    pm = {c["name"]: c for c in prev.get("caps", [])}
    out = []
    for c in cur.get("caps", []):
        p = pm.get(c["name"])
        out.append({"name": c["name"], "new": p is None,
                    "va": (p or {}).get("verdict", "-"), "vb": c["verdict"],
                    "aa": (p or {}).get("access"), "ab": c["access"],
                    "ra": (p or {}).get("rows"), "rb": c["rows"]})
    gone = [n for n in pm if n not in {c["name"] for c in cur.get("caps", [])}]
    for n in gone:
        p = pm[n]
        out.append({"name": n, "new": False, "dropped": True,
                    "va": p["verdict"], "vb": "-", "aa": p["access"], "ab": None,
                    "ra": p["rows"], "rb": None})
    return out


def build_changes_html(path, prev, cur, changes):
    esc = lambda s: (str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))
    n = lambda v: "-" if v is None else f"{v:,}"

    def cell(a, b, fmt=str):
        if a == b:
            return f'<td class="same">{esc(fmt(b))}</td>'
        return (f'<td class="moved"><span class="was">{esc(fmt(a))}</span>'
                f'<span class="arw">&#8594;</span><b>{esc(fmt(b))}</b></td>')

    body = []
    for r in _cmp_rows(prev, cur):
        tag = (' <span class="tg">new</span>' if r.get("new")
               else ' <span class="tg">no longer measured</span>' if r.get("dropped") else "")
        body.append(f'<tr><td class="nm">{esc(r["name"])}{tag}</td>'
                    + cell(r["va"], r["vb"]) + cell(r["aa"], r["ab"], n)
                    + cell(r["ra"], r["rb"], n) + "</tr>")

    det = []
    for name, d in changes["by_name"].items():
        li = "".join(f"<li>{esc(t)}</li>" for t in (d["head"] + d["detail"]))
        det.append(f'<div class="cap2"><b>{esc(name)}</b><ul>{li}</ul></div>')
    if not det:
        det = ['<p class="none">Nothing measured on this map moved between these two runs. '
               'That is a finding in its own right: the idle capabilities are still idle, '
               'and the access that opens them is unchanged.</p>']

    net = sum((r["ab"] or 0) - (r["aa"] or 0) for r in _cmp_rows(prev, cur))
    html = f"""<!doctype html><html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Banner Capability Map: what moved</title><link rel="icon" href="data:,">
<style>
*{{box-sizing:border-box}}
body{{font-family:-apple-system,Segoe UI,Roboto,sans-serif;margin:0;color:#1f2933;
background:#f4f6f8}}
.pg{{max-width:1000px;margin:0 auto;padding:26px 28px 44px}}
h1{{font-size:23px;margin:0;color:#1F3864}}
.sub{{color:#5B6B7C;font-size:14px;margin:4px 0 0}}
.big{{font-size:29px;font-weight:800;color:#B03A2E;margin:14px 0 0}}
.bigl{{color:#5B6B7C;font-size:13px;margin:2px 0 18px}}
h2{{font-size:14px;color:#1F3864;border-bottom:2px solid #1F3864;padding-bottom:3px;
margin:26px 0 10px;text-transform:uppercase;letter-spacing:.5px}}
table{{width:100%;border-collapse:collapse;font-size:12.5px;background:#fff;
border-radius:8px;overflow:hidden;box-shadow:0 1px 4px rgba(0,0,0,.06)}}
th{{background:#1F3864;color:#fff;font-size:10px;text-transform:uppercase;
letter-spacing:.5px;padding:8px 10px;text-align:left}}
td{{padding:7px 10px;border-bottom:1px solid #eef1f4;vertical-align:top}}
td.nm{{font-weight:700}}
td.same{{color:#9aa5b1}}
td.moved{{background:#fff8f2}}
td.moved .was{{color:#9aa5b1;text-decoration:line-through}}
td.moved .arw{{color:#B03A2E;margin:0 5px}}
.tg{{font-size:9.5px;font-weight:700;text-transform:uppercase;letter-spacing:.4px;
background:#eef3fb;color:#1F3864;padding:1px 5px;border-radius:3px}}
.cap2{{background:#fff;border-left:4px solid #1F3864;border-radius:0 6px 6px 0;
padding:10px 14px;margin:0 0 8px;box-shadow:0 1px 3px rgba(0,0,0,.05)}}
.cap2 b{{color:#1F3864;font-size:13.5px}}
.cap2 ul{{margin:5px 0 0;padding-left:18px;font-size:12.5px;line-height:1.6;color:#3a4a5c}}
.none{{background:#f2f9f5;border-left:4px solid #1E6F50;padding:13px 15px;font-size:13px;
color:#2a4a3a;border-radius:0 6px 6px 0;line-height:1.55;margin:0}}
footer{{font-size:11px;color:#9aa5b1;margin-top:30px;line-height:1.6}}
</style></head><body><div class="pg">
<h1>Banner Capability Map: what moved</h1>
<p class="sub">Comparing the run of <b>{esc(prev.get('date','?'))}</b> against the run of
<b>{esc(cur.get('date','?'))}</b>. Read entirely from the two stored runs, so no query
touched Banner to produce this page.</p>
<div class="big">{changes['n']} capabilit{'y' if changes['n'] == 1 else 'ies'} moved</div>
<p class="bigl">Net change in people who can reach a capability across the whole map:
{'+' if net > 0 else ''}{net:,}. People are counted, never listed.</p>
<h2>What moved, capability by capability</h2>
{''.join(det)}
<h2>Every capability, side by side</h2>
<table><thead><tr><th>Capability</th><th>Verdict</th><th>People who can reach it</th>
<th>Records</th></tr></thead><tbody>{''.join(body)}</tbody></table>
<footer>Greyed values did not change between the two runs; they are shown so a stable map
can be told apart from a partial one. Access counts exclude service accounts and
no-access roles. Database evidence shows technical artifacts, not licence entitlement or
functional fit: validate with the Banner product owner before acting.</footer>
</div></body></html>"""
    with open(path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  wrote {path}")


def build_changes_excel(path, prev, cur, changes):
    wb = Workbook(); ws = wb.active; ws.title = "Side by side"
    ws.append([f"Banner Capability Map: {prev.get('date','?')} compared with "
               f"{cur.get('date','?')}"])
    ws["A1"].font = Font(bold=True, size=13, color="1F3864")
    ws.append([])
    ws.append(["Capability", "Verdict then", "Verdict now", "People then", "People now",
               "Records then", "Records now", "Moved"])
    for c in ws[3]:
        c.fill, c.font = HDR, HDRF
    for r in _cmp_rows(prev, cur):
        moved = (r["va"] != r["vb"]) or (r["aa"] != r["ab"]) or (r["ra"] != r["rb"])
        ws.append([r["name"], r["va"], r["vb"], r["aa"], r["ab"],
                   "n/a" if r["ra"] is None else r["ra"],
                   "n/a" if r["rb"] is None else r["rb"], "yes" if moved else ""])
        if moved:
            for c in ws[ws.max_row]:
                c.fill = PatternFill("solid", fgColor="FDF2F0")
    for i, w in enumerate([34, 18, 18, 12, 12, 14, 14, 8], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"
    if changes.get("by_name"):
        _since_sheet(wb, changes)
    wb.save(path)
    print(f"  wrote {path}")


# ---------------------------------------------------------------------------------
#  THE ONE PICTURE.  Every verdict on this map is really one sentence: how many people
#  can open the door, against how much is inside the room. Plotted, the argument makes
#  itself, and the bottom strip (keys handed out, nothing written, ever) needs no
#  caption at all. Drawn server-side as plain SVG so it survives printing and email.
# ---------------------------------------------------------------------------------
CHART_COLOR = {"owned": "#B03A2E", "abandoned": "#B03A2E", "fading": "#C9A227",
               "custom-built": "#C9A227", "in-use": "#1E6F50",
               "not-installed": "#9aa5b1", "expected-absent": "#9aa5b1"}


def _chart_svg(data, esc):
    import math
    pts = [(i, r) for i, r in enumerate(data) if r["rows"] is not None]
    skipped = [r for r in data if r["rows"] is None]
    if not pts:
        return ""
    W, L, R, TOP, BOT, ZERO = 1120, 104, 1066, 46, 292, 348
    xmax = max(50, int(math.ceil(max(r["access"] for _, r in pts) / 50.0) * 50))
    biggest = max((r["rows"] for _, r in pts if r["rows"]), default=10)
    emax = max(1, int(math.ceil(math.log10(biggest))))

    def X(a):
        return L + (a / xmax) * (R - L)

    def Y(n):                                  # log, because 0 to 9.6M on a linear
        if not n:                              # axis would flatten everything but BCM
            return ZERO
        return BOT - (math.log10(n) / emax) * (BOT - TOP)

    def short(nm):
        """Shorten at a boundary that still reads. Cutting mid-word ("Salary Planner
        (budget plan") makes the chart look broken, so drop the parenthetical first,
        then the second half of a slashed name, and only then give up and clip."""
        s = nm.split("(")[0].strip()
        if len(s) > 30:
            s = s.split("/")[0].strip() or s
        return s if len(s) <= 30 else s[:29].rstrip() + "…"

    items = []
    for i, r in pts:
        items.append(dict(i=i, x=X(r["access"]), y=Y(r["rows"]), lbl=short(r["e"]["name"]),
                          full=r["e"]["name"], acc=r["access"], rows=r["rows"],
                          zero=not r["rows"],
                          c=CHART_COLOR.get(verdict_cls(r["verdict"]), "#9aa5b1")))
    #  The zero strip is where the story is, and it is also the most crowded: those
    #  capabilities sit within a few dozen people of each other, so their dots would
    #  print on top of one another. Stagger them inside the strip so each one is
    #  visible, and let the leader lines carry the labels down out of the way.
    for n, it in enumerate(sorted([i for i in items if i["zero"]], key=lambda d: d["x"])):
        it["y"] = ZERO - 9 + 18 * (n % 2)

    #  Label placement. 16 points, several of them stacked in the zero strip, so a
    #  label that simply sits next to its dot would overprint its neighbour. Push down
    #  until clear and draw a leader line whenever the label had to leave its dot.
    placed = []
    dots = [(i["x"], i["y"]) for i in items]
    #  Inside the strip, place in x order: labels then stack the way the dots sit and
    #  the leader lines run parallel instead of crossing each other.
    for it in sorted(items, key=lambda d: (d["zero"], 0 if d["zero"] else d["y"], d["x"])):
        w = 6.05 * len(it["lbl"]) + 8
        #  A label for a dot inside the strip starts BELOW the strip, never on it,
        #  so it cannot print over a neighbouring dot.
        lx, ly = it["x"] + 11, (ZERO + 34 if it["zero"] else it["y"] + 3.6)
        if lx + w > W - 8:                     # would run off the right edge: flip it
            lx = it["x"] - 11 - w
        for _ in range(80):
            hits_label = any(lx < q[0] + q[2] and q[0] < lx + w and abs(ly - q[1]) < 12.5
                             for q in placed)
            #  A label must clear the other DOTS too, not just the other labels. Text
            #  printed across a neighbour's dot is the one thing that makes a chart
            #  look wrong at a glance.
            hits_dot = any(lx - 3 < dx < lx + w + 3 and abs(ly - 3.5 - dy) < 9
                           for dx, dy in dots)
            if not (hits_label or hits_dot):
                break
            ly += 12.5
        placed.append((lx, ly, w))
        it["lx"], it["ly"], it["w"] = lx, ly, w
    #  The people axis goes UNDER the deepest label, not at a fixed height. Otherwise a
    #  crowded zero strip pushes its labels straight through the axis numbers.
    AX = max(ZERO + 40, max(i["ly"] for i in items) + 24)
    H = int(AX + 52)

    o = [f'<svg viewBox="0 0 {W} {H}" class="chart" role="img" '
         f'aria-label="People who can reach each capability against how many records it holds">']
    o.append(f'<rect x="{L - 40}" y="{ZERO - 21}" width="{R - L + 92}" height="42" '
             f'rx="6" fill="#fdf2f0"/>')
    o.append(f'<text x="{L - 46}" y="{ZERO + 4}" class="cax cz" text-anchor="end">'
             f'no records<tspan x="{L - 46}" dy="12">at all</tspan></text>')
    for e in range(emax + 1):                                   # log gridlines
        y = Y(10 ** e)
        lab = "1" if e == 0 else ("10" if e == 1 else
                                  f"{10 ** e:,}" if e < 6 else f"{10 ** (e - 6)}M")
        o.append(f'<line x1="{L}" y1="{y:.1f}" x2="{R}" y2="{y:.1f}" class="cgrid"/>')
        o.append(f'<text x="{L - 8}" y="{y + 3.5:.1f}" class="cax" text-anchor="end">{lab}</text>')
    for t in range(0, 5):                                       # people axis
        a = xmax * t // 4
        x = X(a)
        o.append(f'<line x1="{x:.1f}" y1="{TOP - 12}" x2="{x:.1f}" y2="{ZERO + 21}" class="cgridv"/>')
        o.append(f'<text x="{x:.1f}" y="{AX:.0f}" class="cax" text-anchor="middle">{a}</text>')
    o.append(f'<text x="{(L + R) / 2:.0f}" y="{AX + 22:.0f}" class="cax ct" '
             f'text-anchor="middle">people who can reach the screens &#8594;</text>')
    o.append(f'<text transform="translate(22,{(TOP + BOT) / 2:.0f}) rotate(-90)" '
             f'class="cax ct" text-anchor="middle">records held (log scale) &#8593;</text>')
    for it in items:                                            # leaders, then dots
        right = it["lx"] > it["x"]
        ax = it["lx"] - 4 if right else it["lx"] + it["w"] + 4
        if abs(it["ly"] - (it["y"] + 3.6)) > 3 or not right:
            o.append(f'<line x1="{it["x"]:.1f}" y1="{it["y"]:.1f}" x2="{ax:.1f}" '
                     f'y2="{it["ly"] - 3.5:.1f}" class="clead"/>')
    for it in items:
        anchor = "start" if it["lx"] > it["x"] else "end"
        tx = it["lx"] if anchor == "start" else it["lx"] + it["w"]
        rec = "no records" if not it["rows"] else f'{it["rows"]:,} records'
        o.append(f'<g class="pt" data-i="{it["i"]}"><title>{esc(it["full"])}: '
                 f'{it["acc"]} people can reach it, {rec}</title>'
                 f'<circle cx="{it["x"]:.1f}" cy="{it["y"]:.1f}" r="6.5" fill="{it["c"]}" '
                 f'fill-opacity=".85" stroke="#fff" stroke-width="1.5"/>'
                 f'<text x="{tx:.1f}" y="{it["ly"]:.1f}" class="clbl" '
                 f'text-anchor="{anchor}" fill="{it["c"]}">{esc(it["lbl"])}</text></g>')
    o.append("</svg>")
    note = ("Each dot is one capability: how many people can open its screens, against how "
            "much it holds. The strip along the bottom is the finding. Those are rooms with "
            "doors built and keys handed out, where not one record has ever been written.")
    if skipped:
        note += (" Not shown: " + ", ".join(esc(r["e"]["name"]) for r in skipped) +
                 ", not installed here, so there is nothing to count.")
    return ('<div class="chartbox"><div class="charth">Keys handed out, against what is '
            'inside the room</div>' + "".join(o) +
            f'<p class="chartn">{note} Click a dot to open it.</p></div>')


# ---------------------------------------------------------------------------------
#  INTERACTIVE HTML  (self-contained: click a capability, its detail fills the panel)
# ---------------------------------------------------------------------------------
def build_html(path, data, changes=None, demo=False):
    cls = verdict_cls

    payload = []
    for i, r in enumerate(data):
        e = r["e"]
        payload.append({
            "i": i, "name": e["name"], "buys": e["buys"], "verdict": r["verdict"],
            "why": r["why"], "note": e["note"], "access": r["access"],
            "rows": r["rows"], "last": r["last"], "cls": cls(r["verdict"]),
            "forms": [{"n": n, "d": d} for n, d in r["forms"]],
            "tables": r["tables"],
            "paths": r.get("paths") or [],
            "debt": bool(r.get("debt")),
            "census": r.get("census") or {"total": 0, "with_data": 0},
            "lf_stat": e.get("lf_stat", "Banner screens"),
            "lf_sec": e.get("lf_sec", "Screens that ship in Banner"),
            "sql": method_sql(e),
            #  A capability the boss can send as a link, so the conversation that follows
            #  this report starts on the right card instead of "scroll until you see it".
            "slug": slug(e["name"]),
            "chg": ((changes or {}).get("by_name") or {}).get(e["name"]),
        })

    groups = [
        ("Owned, and sitting idle", ["Owned, not used", "Abandoned"],
         "Banner ships the screens, people hold the keys, the room is empty. Before buying a "
         "tool for any of these, ask why the one you own is dark."),
        ("Owned, but solved another way", ["Custom-built", "Fading", "Owned, data present"],
         "Banner's own module is mostly dark, but the capability is met another way, sometimes "
         "by custom code. Worth a conversation before buying a third product for it."),
        ("Already working (do not re-buy)", ["In use"],
         "A department asking to buy a tool for one of these already has a live one."),
        ("Not part of this Banner", ["Not installed", "Expected absent"],
         "Banner offers these; your institution does not run them. One may be a licensing "
         "choice; housing is simply not what a college without residence halls needs."),
    ]
    esc = lambda s: (str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))
    listhtml = []
    for title, verds, blurb in groups:
        items = [r for r in data if r["verdict"] in verds]
        if not items:
            continue
        rowhtml = []
        for r in items:
            idx = data.index(r)
            e = r["e"]
            rec = "not installed" if r["rows"] is None else (f"{r['rows']:,} records"
                                                             if r["rows"] else "0 records")
            rowhtml.append(
                f'<button class="cap {cls(r["verdict"])}" data-i="{idx}" '
                f'id="c-{slug(e["name"])}">'
                f'<span class="cn">{esc(e["name"])}</span>'
                f'<span class="v">{esc(r["verdict"])}</span>'
                f'<span class="mini">{len(r["forms"])} forms &middot; '
                f'{r["access"]} can reach &middot; {rec}</span></button>')
        listhtml.append(f'<h2>{esc(title)}</h2><p class="blurb">{esc(blurb)}</p>'
                        + "".join(rowhtml))
    owned = sum(1 for r in data if r["verdict"] in ("Owned, not used", "Abandoned"))
    j = json.dumps(payload)

    html = """<!doctype html><html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Banner Capability Map</title>
<link rel="icon" href="data:,">
<style>
*{box-sizing:border-box}
body{font-family:-apple-system,Segoe UI,Roboto,sans-serif;margin:0;color:#1f2933;background:#f4f6f8}
header{padding:22px 28px 10px;max-width:1200px;margin:0 auto}
h1{font-size:24px;margin:0;color:#1F3864}
.sub{color:#5B6B7C;margin:2px 0 0;font-size:14px}
.head{font-size:30px;font-weight:800;color:#B03A2E;margin:12px 0 0}
.headl{color:#5B6B7C;font-size:13px;margin:2px 0 0}
.wrap{display:flex;gap:20px;max-width:1200px;margin:0 auto;padding:12px 28px 40px;align-items:flex-start}
.list{flex:1 1 46%;min-width:340px}
h2{font-size:15px;color:#1F3864;border-bottom:2px solid #1F3864;padding-bottom:3px;margin:20px 0 3px}
.blurb{color:#5B6B7C;font-size:12px;margin:0 0 8px}
.cap{display:block;width:100%;text-align:left;border:0;border-left:5px solid #ccc;background:#fff;
padding:9px 12px;margin:6px 0;border-radius:0 6px 6px 0;cursor:pointer;box-shadow:0 1px 2px rgba(0,0,0,.05)}
.cap:hover{box-shadow:0 2px 8px rgba(0,0,0,.12)}
.cap.sel{outline:2px solid #1F3864;outline-offset:1px}
.cap.owned,.cap.abandoned{border-color:#B03A2E}
.cap.fading,.cap.custom-built{border-color:#C9A227}
.cap.in-use{border-color:#1E6F50}
.cap.not-installed,.cap.expected-absent{border-color:#9aa5b1}
.cn{font-weight:700;font-size:14px}
.v{float:right;font-size:10px;font-weight:700;color:#5B6B7C;text-transform:uppercase;letter-spacing:.4px}
.mini{display:block;font-size:11.5px;color:#5B6B7C;margin-top:2px}
.panel{flex:1 1 54%;position:sticky;top:14px;background:#fff;border-radius:8px;
box-shadow:0 2px 10px rgba(0,0,0,.08);padding:18px 20px;min-height:320px}
.panel .ph{color:#9aa5b1;font-size:14px;text-align:center;margin-top:80px}
.pname{font-size:19px;font-weight:800;color:#1F3864;margin:0}
.pv{display:inline-block;font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.5px;
padding:2px 8px;border-radius:10px;margin:6px 0 2px;color:#fff}
.pv.owned,.pv.abandoned{background:#B03A2E}.pv.fading,.pv.custom-built{background:#C9A227}
.pv.in-use{background:#1E6F50}.pv.not-installed,.pv.expected-absent{background:#9aa5b1}
.pbuys{color:#8A6D3B;font-size:12.5px;margin:6px 0}
.stats{display:flex;gap:10px;margin:12px 0}
.stat{flex:1;background:#f4f6f8;border-radius:6px;padding:8px 10px;text-align:center}
.stat b{display:block;font-size:20px;color:#1F3864}
.stat span{font-size:10.5px;color:#5B6B7C}
.sec{font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.5px;color:#5B6B7C;margin:14px 0 4px}
.tbl{width:100%;border-collapse:collapse;font-size:12px}
.tbl td{padding:3px 6px;border-bottom:1px solid #eef1f4;vertical-align:top}
.tbl td.n{color:#5B6B7C;text-align:right;white-space:nowrap}
.tbl td.td{color:#3a4a5c;font-size:11.5px}
.tbl th{padding:3px 6px;border-bottom:2px solid #d7dde4;text-align:left;font-size:9.5px;font-weight:700;text-transform:uppercase;letter-spacing:.4px;color:#8a97a6;white-space:nowrap}
.tbl th.n{text-align:right}
.tbl tr.pri td{background:#eef3fb}
.census{font-size:12px;color:#8a3a2e;background:#fbf1ee;border-radius:4px;padding:6px 9px;margin:2px 0 6px}
.judge{font-size:9.5px;font-weight:700;text-transform:uppercase;letter-spacing:.4px;color:#1F3864;
background:#dbe6f7;padding:1px 5px;border-radius:8px;margin-left:6px}
.mono{font-family:ui-monospace,Consolas,monospace;font-weight:600}
.note{background:#fbfaf3;border-left:3px solid #C9A227;padding:8px 10px;font-size:12px;color:#5B6B7C;margin-top:12px;border-radius:0 4px 4px 0}
/*  the group layer: a Banner GROUP bundles classes, a CLASS bundles screens, and a
    direct grant is pinned onto one person. Colour says which, so the eye sorts it. */
.paths .kind{font-size:9.5px;font-weight:700;text-transform:uppercase;letter-spacing:.4px;
padding:1px 5px;border-radius:3px;white-space:nowrap}
.paths .k-group .kind{background:#e4ecf7;color:#1F3864}
.paths .k-class .kind{background:#eef2f5;color:#5B6B7C}
.paths .k-direct .kind{background:#fbe9e6;color:#B03A2E}
.paths .k-direct .mono{color:#B03A2E}
.paths .gdesc{display:block;font-size:12.5px;color:#1f2933}
.paths .gcode{display:block;font-family:ui-monospace,Consolas,monospace;font-size:10.5px;color:#9aa5b1}
.paths .chg{color:#B03A2E}
.paths .zero{color:#b7c0c9}
.debt{background:#fdf2f0;border-left:3px solid #B03A2E;padding:8px 10px;font-size:12px;
color:#8a3a2e;margin:8px 0 0;border-radius:0 4px 4px 0}
/*  the one picture, and the band that says what moved since the last run  */
.chartbox{max-width:1200px;margin:14px auto 0;padding:16px 28px 6px;background:#fff;
border-radius:8px;box-shadow:0 2px 10px rgba(0,0,0,.06)}
.charth{font-size:15px;font-weight:800;color:#1F3864;margin:0 0 2px}
.chart{width:100%;height:auto;display:block;overflow:visible}
.chart .cgrid{stroke:#eef1f4;stroke-width:1}
.chart .cgridv{stroke:#f4f6f8;stroke-width:1}
.chart .cax{font:10.5px -apple-system,Segoe UI,Roboto,sans-serif;fill:#9aa5b1}
.chart .cax.ct{font-size:11px;fill:#5B6B7C;font-weight:600}
.chart .cax.cz{fill:#B03A2E;font-weight:700}
.chart .clbl{font:10.5px -apple-system,Segoe UI,Roboto,sans-serif;font-weight:600}
.chart .clead{stroke:#c3ccd6;stroke-width:.8}
.chart .pt{cursor:pointer}
.chart .pt:hover circle{r:9}
.chart .pt:hover .clbl{text-decoration:underline}
.chartn{font-size:12px;color:#5B6B7C;margin:2px 0 8px;line-height:1.5;max-width:960px}
.since{max-width:1200px;margin:14px auto 0;padding:12px 18px;background:#eef3fb;
border-left:4px solid #1F3864;border-radius:0 6px 6px 0;font-size:13px;color:#2a3a52}
.since b.st{display:block;font-size:11px;font-weight:800;text-transform:uppercase;
letter-spacing:.5px;color:#1F3864;margin-bottom:5px}
.since ul{margin:0;padding-left:18px;line-height:1.6}
.since .cn2{font-weight:700}
.since .more{color:#5B6B7C;font-size:12px;margin:5px 0 0}
.chgbox{background:#eef3fb;border-left:3px solid #1F3864;padding:8px 10px;font-size:12px;
color:#2a3a52;margin:10px 0 0;border-radius:0 4px 4px 0}
.chgbox b.ct2{display:block;font-size:10px;font-weight:800;text-transform:uppercase;
letter-spacing:.4px;color:#1F3864;margin-bottom:4px}
.chgbox ul{margin:0;padding-left:16px;line-height:1.55}
/*  A demo copy must be impossible to mistake for the real report, including by
    someone who walks past a screen or is handed a printout.  */
.demo{background:#B03A2E;color:#fff;padding:11px 28px;font-size:13px;line-height:1.5}
.demo b{display:block;font-size:14px;letter-spacing:.4px;text-transform:uppercase}
@media print{.demo{-webkit-print-color-adjust:exact;print-color-adjust:exact}}
footer{max-width:1200px;margin:0 auto;padding:0 28px 30px;font-size:11px;color:#9aa5b1}
@media(max-width:820px){.wrap{flex-direction:column}.panel{position:static;width:100%}}
.how{display:inline-block;margin-top:12px;background:#1F3864;color:#fff;border:0;border-radius:20px;
padding:9px 18px;font-size:13px;font-weight:600;cursor:pointer}
.how:hover{background:#2a4a7f}
.modal{display:none;position:fixed;inset:0;background:rgba(20,30,45,.55);z-index:50;overflow:auto}
.modal.open{display:block}
.sheet{max-width:880px;margin:26px auto;background:#fff;border-radius:12px;padding:28px 34px 38px;
position:relative;box-shadow:0 20px 60px rgba(0,0,0,.35)}
.x{position:absolute;top:12px;right:16px;border:0;background:transparent;font-size:28px;color:#9aa5b1;cursor:pointer;line-height:1}
.mt{font-size:25px;color:#1F3864;margin:0 0 4px}
.mlead{color:#5B6B7C;font-size:14px;margin:0;line-height:1.55}
.mstep{font-size:12px;font-weight:800;text-transform:uppercase;letter-spacing:.6px;color:#B03A2E;
margin:26px 0 10px;border-top:2px solid #f1e7e4;padding-top:15px}
.rooms{display:flex;gap:12px;flex-wrap:wrap}
.room{flex:1;min-width:200px;border-radius:9px;padding:14px 16px;border:1px solid #eef1f4}
.room .rt{font-size:26px;letter-spacing:3px;margin-bottom:8px}
.room .rl{font-weight:800;font-size:13px;color:#1F3864}
.room .rd{font-size:12px;color:#5B6B7C;margin-top:3px;line-height:1.45}
.room.idle{background:#fdf4f2;border-color:#f2c9bf}
.room.live{background:#f2f9f5;border-color:#bfe0cd}
.room.none{background:#f7f8f9;border-color:#dde2e8}
.q3{display:flex;gap:12px;flex-wrap:wrap}
.qcard{flex:1;min-width:210px;background:#fafbfc;border-radius:9px;padding:15px;border-top:3px solid #1F3864}
.qcard svg{width:26px;height:26px;margin-bottom:7px}
.qcard b{display:block;font-size:14px;color:#1F3864}
.qcard span{display:block;font-size:12px;color:#5B6B7C;margin-top:4px;line-height:1.5}
.qcard .rd{font-size:10.5px;color:#9aa5b1;margin-top:6px;font-family:ui-monospace,Consolas,monospace}
.adown{text-align:center;color:#c3ccd6;font-size:22px;margin:10px 0 4px}
.verd{display:flex;flex-direction:column;gap:0}
.vrow{display:flex;align-items:center;gap:12px;font-size:13px;color:#3a4a5c;padding:9px 2px;border-bottom:1px solid #eef1f4}
.vrow .cond{flex:1}
.chip{font-size:11px;font-weight:800;text-transform:uppercase;letter-spacing:.4px;color:#fff;
padding:3px 10px;border-radius:11px;white-space:nowrap}
.chip.red{background:#B03A2E}.chip.amber{background:#C9A227}.chip.green{background:#1E6F50}.chip.grey{background:#9aa5b1}
.warn{background:#fbf7ee;border-left:4px solid #C9A227;padding:13px 15px;font-size:13px;color:#5a4a2a;border-radius:0 6px 6px 0;line-height:1.55}
.rule{background:#eef3fb;border-left:4px solid #1F3864;padding:13px 15px;font-size:13px;color:#2a3a52;border-radius:0 6px 6px 0;line-height:1.55}
.sqlrow{display:flex;align-items:center;gap:7px;flex-wrap:wrap;margin:11px 0 2px;font-size:11px;color:#8a97a6}
.sqlc{cursor:pointer;border:1px solid #cfe0f0;background:#eef5fb;color:#1F3864;font:700 11px/1 inherit;
letter-spacing:.3px;padding:5px 9px;border-radius:6px}
.sqlc:hover{background:#dcebf8}
.sqlc::before{content:"</> ";font-family:Consolas,monospace;font-size:10px;opacity:.85}
.sqlback{position:fixed;inset:0;background:rgba(20,30,45,.45);opacity:0;visibility:hidden;transition:.2s;z-index:60}
.sqlback.open{opacity:1;visibility:visible}
.sqldraw{position:fixed;top:0;right:0;height:100vh;width:min(600px,94vw);background:#141a22;color:#e6edf3;
box-shadow:-8px 0 30px rgba(0,0,0,.35);transform:translateX(100%);transition:transform .26s cubic-bezier(.4,0,.2,1);
z-index:70;display:flex;flex-direction:column}
.sqldraw.open{transform:translateX(0)}
.sqlh{display:flex;align-items:center;justify-content:space-between;gap:12px;padding:18px 20px;border-bottom:1px solid #263041}
.sqlh .t{font-size:14px;font-weight:700;color:#fff}
.sqlh .t .qn{display:block;font-size:11px;font-weight:600;color:#7fb2e6;text-transform:uppercase;letter-spacing:.5px;margin-bottom:3px}
.sqlxx{cursor:pointer;background:none;border:0;color:#8a97a6;font-size:24px;line-height:1;padding:0 4px}
.sqlxx:hover{color:#fff}
.sqlbd{padding:18px 20px;overflow-y:auto;flex:1}
.sqlbd .why{font-size:13px;color:#9fb0c2;margin:0 0 14px;line-height:1.5}
.sqlbd pre{margin:0;background:#0d1117;border:1px solid #263041;border-radius:10px;padding:16px 18px;
overflow-x:auto;font:12.5px/1.5 Consolas,Menlo,monospace;color:#e6edf3;white-space:pre}
.sqlbd pre .k{color:#7fb2e6}.sqlbd pre .c{color:#6b7683;font-style:italic}
.sqlbar{display:flex;gap:10px;padding:14px 20px;border-top:1px solid #263041;align-items:center}
.sqlcopy{cursor:pointer;background:#1f6feb;border:0;color:#fff;font:700 12px/1 inherit;padding:9px 16px;border-radius:7px}
.sqlcopy:hover{background:#388bfd}.sqlcopy.done{background:#2ea043}
.sqlnote{font-size:11.5px;color:#6b7683}
@media print{.sqlrow{display:none}}
</style></head><body>
__DEMO__
<header>
<h1>Banner Capability Map</h1>
<p class="sub">What your institution already owns in Banner, and whether it is used.</p>
<div class="head">__OWNED__ capabilities owned and idle</div>
<p class="headl">Measured on the read-only copy of Banner on __DATE__. Click any capability for its
screens, its tables, its records, and the security groups that open it. The drill-down stops at the
group: no individual is named anywhere in this report. Each capability keeps its own link in the
address bar, so a single card can be sent on its own.</p>
<button class="how" onclick="openHow()">&#9654;&nbsp; How we got here: the method behind the map</button>
</header>
__SINCE__
__CHART__
<div class="wrap">
<div class="list">__LIST__</div>
<div class="panel" id="panel"><div class="ph">Click a capability on the left to see its screens,
its tables, and the numbers behind the verdict.</div></div>
</div>
<div id="how" class="modal" onclick="if(event.target===this)closeHow()">
 <div class="sheet">
  <button class="x" onclick="closeHow()" title="close">&times;</button>
  <h2 class="mt">How we got here</h2>
  <p class="mlead">No opinions. Every verdict on this map comes from three questions we ask Banner
  itself, and Banner answers with its own data. Here is the whole method, in plain terms.</p>

  <div class="mstep">1 &middot; The idea, in one picture</div>
  <p style="font-size:13.5px;color:#3a4a5c;margin:0 0 12px;line-height:1.55">A Banner capability is a
  <b>room</b>. The screens Banner ships are the <b>doors</b>. The people with access are who holds a
  <b>key</b>. The data inside is whether anyone actually walks in and uses it.</p>
  <div class="rooms">
    <div class="room idle"><div class="rt">&#128682; &#128273; &nbsp;&#183;&#183;&#183;</div>
      <div class="rl">Owned, and idle</div>
      <div class="rd">Doors built, keys handed out, nobody inside. This is the shelfware, and the purchase to question.</div></div>
    <div class="room live"><div class="rt">&#128682; &#128273; &#128101;&#128101;</div>
      <div class="rl">In use</div>
      <div class="rd">Doors, keys, and people working inside. A live tool already exists. Do not re-buy.</div></div>
    <div class="room none"><div class="rt" style="opacity:.4">&#9634; &#9634; &#9634;</div>
      <div class="rl">Not installed</div>
      <div class="rd">Banner offers it, but the room was never built here. Maybe a licensing choice.</div></div>
  </div>

  <div class="mstep">2 &middot; The three questions we ask Banner</div>
  <div class="q3">
    <div class="qcard">
      <svg viewBox="0 0 24 24"><rect x="5" y="3" width="13" height="18" rx="1" fill="none" stroke="#1F3864" stroke-width="2"/><circle cx="14.5" cy="12" r="1.2" fill="#1F3864"/></svg>
      <b>Do the screens exist?</b>
      <span>If the forms ship, the capability is installed here, not an orphan table.</span>
      <div class="rd">reads GUBOBJS</div></div>
    <div class="qcard">
      <svg viewBox="0 0 24 24"><circle cx="8" cy="8" r="4" fill="none" stroke="#1F3864" stroke-width="2"/><path d="M10.8 10.8 L20 20 M17 17 L19.2 14.8 M20 20 L17.8 22.2" stroke="#1F3864" stroke-width="2" fill="none" stroke-linecap="round"/></svg>
      <b>Can anyone reach them?</b>
      <span>Who holds access, from the same security data behind our access review. The open door, not the traffic.</span>
      <div class="rd">reads GUVUACC</div></div>
    <div class="qcard">
      <svg viewBox="0 0 24 24"><path d="M4 20 V5 M4 20 H21" stroke="#1F3864" stroke-width="2" fill="none" stroke-linecap="round"/><rect x="7" y="12" width="3" height="6" fill="#1F3864"/><rect x="12" y="9" width="3" height="9" fill="#1F3864"/><rect x="17" y="14" width="3" height="4" fill="#1F3864"/></svg>
      <b>Is there data, and recent?</b>
      <span>A row count and the newest activity date on the capability's main table. This is the judge.</span>
      <div class="rd">COUNT + last activity date</div></div>
  </div>

  <div class="mstep">3 &middot; How the verdict is decided</div>
  <div class="adown">&#9660;</div>
  <div class="verd">
    <div class="vrow"><span class="cond">The main table is <b>not there</b> at all</span><span class="chip grey">Not installed</span></div>
    <div class="vrow"><span class="cond">The table Banner ships for this is <b>empty</b> (confirm no self-service writes elsewhere)</span><span class="chip red">Owned, not used</span></div>
    <div class="vrow"><span class="cond">It is empty, but a <b>custom table</b> quietly does the job</span><span class="chip amber">Custom-built</span></div>
    <div class="vrow"><span class="cond">It is empty <b>by design</b> (housing at a college with no residence halls)</span><span class="chip grey">Expected absent</span></div>
    <div class="vrow"><span class="cond">It has data, but <b>nothing recent</b></span><span class="chip red">Abandoned</span></div>
    <div class="vrow"><span class="cond">It has <b>recent</b> activity</span><span class="chip green">In use</span></div>
    <div class="vrow"><span class="cond">The table is there but <b>could not be read</b> (timeout or privilege)</span><span class="chip amber">Unable to verify</span></div>
  </div>

  <div class="mstep">A wrong turn we deliberately avoided</div>
  <div class="warn">We could have asked "is there an Argos <b>report</b> on it?" But a capability can be used
  entirely inside Banner, through screens and native output that satisfy the department, and never have a
  report built. A missing report proves nothing. So we ask the data and the screens, never the report list.</div>

  <div class="mstep">The one rule that keeps us honest</div>
  <div class="rule">Database evidence shows technical artifacts and rows visible to a read-only account. It does
  <b>not</b> prove license ownership, deployment, adoption, usability, or functional fit. Empty is not the same
  as <b>unlicensed</b>; a busy activity date is not the same as human use. Every "owned, not used" here is a
  signal to validate with the Banner product owner, never a verdict on its own. We show the door and the
  traffic; the contract and the fit are separate questions.</div>
 </div>
</div>
<div class="sqlback" id="sqlback" onclick="closeSql()"></div>
<aside class="sqldraw" id="sqldraw">
  <div class="sqlh"><div class="t"><span class="qn" id="sqlqn">query</span><span id="sqltitle">SQL</span></div>
   <button class="sqlxx" onclick="closeSql()" title="close (Esc)">&times;</button></div>
  <div class="sqlbd"><p class="why" id="sqlwhy"></p><pre id="sqlpre"></pre></div>
  <div class="sqlbar"><button class="sqlcopy" id="sqlcopy" onclick="copySql()">Copy SQL</button>
   <span class="sqlnote">read-only &middot; SELECT only &middot; run it live in SQL Developer</span></div>
</aside>
<footer>Source: a read-only Banner database. Table row counts in the panel are estimates from Banner's own
statistics; the headline record count and last-activity date are exact. This measures visible use, not
entitlement or fit: empty is not unlicensed, and a recent activity date can come from a load job, not a
person. Validate license and workflow with the product owner before acting. Group and class counts read
the security view GUVUACC, which names the vehicle on every grant; people are counted, never listed.
Run by __USER__.</footer>
<script>
var DATA = __JSON__;
function openHow(){document.getElementById('how').classList.add('open');}
function closeHow(){document.getElementById('how').classList.remove('open');}
var SQLKW=/\\b(SELECT|FROM|WHERE|AND|OR|JOIN|ON|GROUP BY|ORDER BY|COUNT|DISTINCT|CASE|WHEN|THEN|ELSE|END|AS|IS|NOT|NULL|LIKE|IN|MAX|FETCH|FIRST|ROWS|ONLY|NULLS|LAST|UPPER|ESCAPE)\\b/g;
function sqlHi(s){return esc(s).replace(/(--.*)/g,'<span class="c">$1</span>').replace(SQLKW,'<span class="k">$&</span>');}
var _cursql='';
function openSql(i,key){
  var d=DATA[i]; if(!d||!d.sql||!d.sql[key])return;
  var s=d.sql[key];
  document.getElementById('sqlqn').textContent=d.name;
  document.getElementById('sqltitle').textContent=s.t;
  document.getElementById('sqlwhy').textContent=s.w;
  document.getElementById('sqlpre').innerHTML=sqlHi(s.q);
  _cursql=s.q;
  var c=document.getElementById('sqlcopy');c.textContent='Copy SQL';c.classList.remove('done');
  document.getElementById('sqldraw').classList.add('open');
  document.getElementById('sqlback').classList.add('open');
}
function closeSql(){
  document.getElementById('sqldraw').classList.remove('open');
  document.getElementById('sqlback').classList.remove('open');
}
function copySql(){
  if(!navigator.clipboard)return;
  navigator.clipboard.writeText(_cursql).then(function(){
    var c=document.getElementById('sqlcopy');c.textContent='Copied';c.classList.add('done');});
}
document.addEventListener('keydown',function(e){if(e.key==='Escape'){closeHow();closeSql();}});
function esc(s){return String(s==null?'':s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');}
function nf(n){return n==null?'?':n.toLocaleString();}
function render(i){
  var d=DATA[i], p=document.getElementById('panel');
  var recs = d.rows==null?'not installed':(nf(d.rows)+(d.rows?' records':' records — empty'));
  var last = d.last?(' &middot; last activity '+esc(d.last)):'';
  var forms = d.forms.length? d.forms.map(function(f){
      return '<tr><td class="mono">'+esc(f.n)+'</td><td>'+esc(f.d)+'</td></tr>';}).join('')
      : '<tr><td colspan=2 style="color:#9aa5b1">no forms found</td></tr>';
  var tabs = d.tables.length? d.tables.map(function(t){
      var tag = t.primary? ' <span class="judge">judges this</span>':'';
      var rc = t.primary? '<b>'+nf(t.rows)+'</b>' : nf(t.rows);
      return '<tr'+(t.primary?' class="pri"':'')+'><td class="mono">'+esc(t.t)+tag+
             '</td><td class="td">'+esc(t.desc||'')+'</td><td class="n">'+rc+
             '</td><td class="n">'+esc(t.analyzed||'-')+'</td></tr>';}).join('')
      : '<tr><td colspan=4 style="color:#9aa5b1">table not present</td></tr>';
  var sqlrow = d.sql ? ('<div class="sqlrow">Show the SQL:'+
      '<button class="sqlc" onclick="openSql('+i+',&#39;screens&#39;)">screens</button>'+
      '<button class="sqlc" onclick="openSql('+i+',&#39;access&#39;)">access</button>'+
      (d.sql.groups?'<button class="sqlc" onclick="openSql('+i+',&#39;groups&#39;)">groups</button>':'')+
      '<button class="sqlc" onclick="openSql('+i+',&#39;records&#39;)">records</button>'+
      (d.sql.tables?'<button class="sqlc" onclick="openSql('+i+',&#39;tables&#39;)">tables</button>':'')+
    '</div>') : '';

  //  THE GROUP LAYER. The map answers "is it used"; this answers "who is it wired to",
  //  and it deliberately stops at the group. No individual is ever named here.
  var ngrp = 0, ndir = 0, k;
  for(k=0;k<d.paths.length;k++){
    if(d.paths[k].kind==='Group') ngrp++;
    if(d.paths[k].kind==='Direct to person') ndir++;
  }
  var paths = d.paths.length? d.paths.map(function(g){
      var chg = g.change? '<b class="chg">'+g.change+'</b>' : '<span class="zero">0</span>';
      //  Banner names groups after the job, so the description IS the readable label.
      //  Lead with it and keep the code underneath as the technical reference.
      var label = g.desc? '<b class="gdesc">'+esc(g.desc)+'</b><span class="gcode">'+
                          esc(g.name)+'</span>'
                        : '<span class="mono">'+esc(g.name)+'</span>';
      return '<tr class="k-'+g.kind.split(' ')[0].toLowerCase()+'">'+
             '<td><span class="kind">'+esc(g.kind)+'</span></td>'+
             '<td>'+label+'</td>'+
             '<td class="n">'+nf(g.screens)+'</td>'+
             '<td class="n">'+chg+'</td>'+
             '<td class="n">'+nf(g.people)+'</td></tr>';}).join('')
      : '<tr><td colspan=5 style="color:#9aa5b1">no security group or class reaches this '+
        'capability (Banner Workflow carries its own user list, separately)</td></tr>';
  //  What moved on THIS capability since the previous run, if there is one to compare.
  var chg='';
  if(d.chg && ((d.chg.head&&d.chg.head.length)||(d.chg.detail&&d.chg.detail.length))){
    var all=(d.chg.head||[]).concat(d.chg.detail||[]);
    chg='<div class="chgbox"><b class="ct2">Since the previous run</b><ul>'+
        all.map(function(t){return '<li>'+esc(t)+'</li>';}).join('')+'</ul></div>';
  }
  var debtline = d.debt? '<div class="debt"><b>Standing access on a dark capability.</b> '+
      'Nothing here is misconfigured. But this capability holds no live data, and a group '+
      'above still grants the right to change it. That is the cheapest cleanup on this page: '+
      'it costs nothing to remove and it is access nobody is using.</div>' : '';
  var pathsum = d.paths.length? '<div class="census">'+
      '<b>'+ngrp+'</b> security group'+(ngrp===1?'':'s')+' and <b>'+
      (d.paths.length-ngrp-ndir)+'</b> stand-alone class'+
      ((d.paths.length-ngrp-ndir)===1?'':'es')+' open this capability'+
      (ndir? ', plus <b>'+ndir+'</b> grant'+(ndir===1?'':'s')+' pinned straight onto a person '+
             '(access no group governs)' : '')+'.</div>' : '';
  p.innerHTML =
    '<p class="pname">'+esc(d.name)+'</p>'+
    '<span class="pv '+d.cls+'">'+esc(d.verdict)+'</span>'+
    '<div class="pbuys">Departments buy for this: '+esc(d.buys)+'</div>'+
    chg+
    '<div class="stats">'+
      '<div class="stat"><b>'+d.forms.length+'</b><span>'+esc(d.lf_stat)+'</span></div>'+
      '<div class="stat"><b>'+nf(d.access)+'</b><span>people can reach them</span></div>'+
      '<div class="stat"><b>'+ngrp+'</b><span>security groups</span></div>'+
      '<div class="stat"><b>'+(d.rows==null?'0':nf(d.rows))+'</b><span>records'+last+'</span></div>'+
    '</div>'+
    sqlrow+
    '<div class="sec">How the access is wired: groups and classes</div>'+
    pathsum+
    '<table class="tbl paths"><thead><tr><th>Kind</th><th>Group or class</th>'+
        '<th class="n">Screens</th><th class="n">Can change</th>'+
        '<th class="n">People</th></tr></thead><tbody>'+paths+'</tbody></table>'+
    debtline+
    '<div class="sec">'+esc(d.lf_sec)+'</div>'+
    '<table class="tbl"><thead><tr><th>Form</th><th>Screen name</th></tr></thead><tbody>'+forms+'</tbody></table>'+
    '<div class="sec">Associated tables</div>'+
    (d.census && d.census.total>1 ? '<div class="census"><b>'+d.census.with_data+
        '</b> of <b>'+d.census.total+'</b> tables in this module hold any data.'+
        (d.tables.length<d.census.total?' Showing '+d.tables.length+' below.':'')+'</div>':'')+
    '<table class="tbl"><thead><tr><th>Table</th><th>What it holds</th>'+
        '<th class="n">Rows (est.)</th><th class="n">Last analyzed</th></tr></thead><tbody>'+
        tabs+'</tbody></table>'+
    '<div class="note">'+esc(d.note)+'</div>';
  var b=document.querySelectorAll('.cap');
  for(var k=0;k<b.length;k++) b[k].classList.toggle('sel', +b[k].dataset.i===i);
}
//  Open a capability AND put it in the address bar, so this report can be sent as a
//  link to one card instead of "open the file and scroll until you find it".
var _quiet=false;
function open_cap(i,scroll){
  render(i);
  _quiet=true; location.hash=DATA[i].slug; _quiet=false;
  if(scroll){var el=document.getElementById('c-'+DATA[i].slug);
             if(el) el.scrollIntoView({block:'center'});}
}
function from_hash(scroll){
  var h=(location.hash||'').replace(/^#/,'');
  if(!h) return false;
  for(var k=0;k<DATA.length;k++) if(DATA[k].slug===h){
    render(k);
    if(scroll){var el=document.getElementById('c-'+h); if(el) el.scrollIntoView({block:'center'});}
    return true;
  }
  return false;
}
document.querySelectorAll('.cap').forEach(function(btn){
  btn.addEventListener('click',function(){open_cap(+this.dataset.i,false);});
});
document.querySelectorAll('.chart .pt').forEach(function(g){
  g.addEventListener('click',function(){open_cap(+this.dataset.i,true);});
});
window.addEventListener('hashchange',function(){if(!_quiet) from_hash(true);});
from_hash(true);
</script>
</body></html>"""
    #  What moved since the previous run. Absent on the very first run, and that absence
    #  is honest: there is nothing to compare against yet.
    since = ""
    if changes and changes.get("head"):
        li = "".join(f'<li><span class="cn2">{esc(n)}</span>: {esc(t)}</li>'
                     for n, t in changes["head"][:8])
        extra = len(changes["head"]) - 8
        more = (f'<p class="more">and {extra} more, on the cards below.</p>'
                if extra > 0 else "")
        since = (f'<div class="since"><b class="st">What moved since '
                 f'{esc(changes["prev_date"])}</b><ul>{li}</ul>{more}</div>')
    elif changes:
        since = (f'<div class="since"><b class="st">Compared against '
                 f'{esc(changes["prev_date"])}</b>Nothing measured on this map has changed '
                 f'since the previous run. The map is stable, which is itself a finding: '
                 f'the idle capabilities are still idle.</div>')

    ribbon = ""
    if demo:
        pd = esc((changes or {}).get("prev_date", "an earlier date"))
        ribbon = (f'<div class="demo"><b>Demonstration copy, do not present</b>'
                  f'The capability measurements below are real and current. The '
                  f'&ldquo;{pd}&rdquo; run they are compared against is INVENTED, so every '
                  f'number describing a change is fictional. It exists only to show what '
                  f'this report will look like once a second real run has been stored.</div>')

    html = (html.replace("__DEMO__", ribbon)
                .replace("__OWNED__", str(owned))
                .replace("__DATE__", datetime.now().strftime("%Y-%m-%d"))
                .replace("__USER__", esc(os.getenv("USERNAME") or "?"))
                .replace("__SINCE__", since)
                .replace("__CHART__", _chart_svg(data, esc))
                .replace("__LIST__", "".join(listhtml))
                .replace("__JSON__", j))
    _check_script(html)
    with open(path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  wrote {path}")


def _check_script(html):
    """The panel's JS lives inside a normal Python string, so a backslash escape written
    for JavaScript (\\n, \\b) is eaten by Python before it ever reaches the browser. That
    turns a regex literal into a syntax error and kills EVERY click on the page at once,
    silently. Catch it here rather than in front of an audience."""
    a = html.index("<script>") + len("<script>")
    js = html[a:html.index("</script>", a)]
    for ch, was in ((chr(8), "\\b"), (chr(12), "\\f"), (chr(11), "\\v"), (chr(7), "\\a")):
        if ch in js:
            sys.exit(f"  BUG: a control character reached the script; a JS '{was}' was "
                     f"eaten by Python. Double the backslash in the HTML template.")
    for ln in js.splitlines():                     # a regex literal cannot span lines
        if ln.count("/") and re.search(r"[=(,]\s*/(?:[^/\\\n]|\\.)*$", ln):
            sys.exit(f"  BUG: unterminated regex literal in the panel JS:\n    {ln.strip()}")
    if shutil.which("node"):                       # free full parse when node is around
        with tempfile.NamedTemporaryFile("w", suffix=".js", encoding="utf-8",
                                         delete=False) as t:
            t.write(js)
            tmp = t.name
        r = subprocess.run(["node", "--check", tmp], capture_output=True, text=True)
        os.unlink(tmp)
        if r.returncode:
            sys.exit("  BUG: the panel JS does not parse, no click would work:\n"
                     + r.stderr.strip())


def main():
    ap = argparse.ArgumentParser(description="Banner capability map -> Excel + presentation.")
    ap.add_argument("--out", default="capability_map.xlsx", help="output .xlsx")
    ap.add_argument("--html", help="also write the interactive one-page presentation here")
    ap.add_argument("--history", help="folder of past runs (default: history/ beside --out)")
    ap.add_argument("--no-save", action="store_true",
                    help="do not add this run to the history")
    ap.add_argument("--since", metavar="DATE",
                    help="compare today's run against the stored run of DATE "
                         "(YYYY-MM-DD; the nearest earlier run is used if that exact "
                         "date was never run)")
    ap.add_argument("--compare", nargs=2, metavar=("FROM", "TO"),
                    help="compare two stored runs and write only the change report. "
                         "Reads the history files, never touches Banner.")
    ap.add_argument("--runs", action="store_true", help="list the stored runs and exit")
    ap.add_argument("--demo", action="store_true",
                    help="stamp the HTML as a demonstration copy. Use whenever the "
                         "baseline being compared against is not a real stored run.")
    ap.add_argument("--snapshot-only", action="store_true",
                    help="measure and save the history snapshot, write no Excel and no "
                         "HTML. This is what the scheduled monthly run uses.")
    a = ap.parse_args()

    hist0 = a.history or os.path.join(os.path.dirname(os.path.abspath(a.out)), "history")
    if a.runs:
        runs = list_runs(hist0)
        print("\n  Stored runs in " + hist0)
        for d in runs:
            print("    " + d)
        print(f"\n  {len(runs)} run(s). Compare any two with:  "
              f"--compare {runs[0] if runs else 'FROM'} "
              f"{runs[-1] if runs else 'TO'}\n")
        return

    #  TWO STORED DATES. No connection, no cursor, no query. This is the mode that runs
    #  in a meeting, off VPN, in a second, and that cannot possibly disturb Banner.
    if a.compare:
        d1, d2 = a.compare
        p, c = load_run(hist0, d1), load_run(hist0, d2)
        if p["date"] == c["date"]:
            sys.exit(f"  Both dates resolve to the same run ({p['date']}). "
                     f"Nothing to compare.")
        if p["date"] > c["date"]:
            p, c = c, p                       # accept the two dates in either order
        changes = diff_snapshots(p, c)
        print(f"\n  Banner capability map: {p['date']} compared with {c['date']}")
        print("  " + "-" * 58)
        if d1 != p["date"] or d2 != c["date"]:
            print(f"  (nearest stored runs used: {p['date']} and {c['date']})")
        print(f"  {changes['n']} capabilities moved")
        for nm, t in changes["head"]:
            print(f"    {nm}: {t}")
        stem = f"changes_{p['date']}_to_{c['date']}"
        base = os.path.dirname(os.path.abspath(a.out))
        build_changes_html(a.html or os.path.join(base, stem + ".html"), p, c, changes)
        build_changes_excel(os.path.join(base, stem + ".xlsx"), p, c, changes)
        print()
        return

    print("\n  Banner capability map")
    print("  " + "-" * 58)
    con = connect(); cur = con.cursor()
    data = build_rows(cur)
    con.close()

    #  HISTORY. Compare against the previous run BEFORE saving this one, otherwise this
    #  run becomes its own baseline and the map can never report that anything moved.
    hist = hist0
    today = datetime.now().strftime("%Y-%m-%d")
    snap = snapshot(data, today)
    #  --since names the baseline explicitly ("what has moved since the March review");
    #  without it the baseline is simply the previous run.
    prev = load_run(hist, a.since) if a.since else load_prev(hist, today)
    if prev is not None and prev.get("date") == today:
        prev = None                       # today cannot be its own baseline
    changes = diff_snapshots(prev, snap) if prev else None
    if changes:
        print(f"\n  since {changes['prev_date']}: {changes['n']} capabilit"
              f"{'y' if changes['n'] == 1 else 'ies'} moved")
        for n, t in changes["head"][:6]:
            print(f"    {n}: {t}")
    elif prev is None:
        print("\n  first run: nothing to compare against yet. Saving the baseline, so "
              "the next run can show what moved.")

    if not a.snapshot_only:
        build_excel(a.out, data, changes, demo=a.demo)
        if a.html:
            build_html(a.html, data, changes, demo=a.demo)
    if not a.no_save:
        save_snapshot(snap, hist)
    owned = sum(1 for r in data if r["verdict"] in ("Owned, not used", "Abandoned"))
    print(f"\n  {owned} capabilities owned and idle. Each is a purchase to question.\n")


if __name__ == "__main__":
    main()
